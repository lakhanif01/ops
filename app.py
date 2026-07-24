"""
OPS — DC Shipping Forecast  (backend)
FastAPI on Hugging Face Spaces · Supabase/PostgREST schema `ops` · memory demo mode when no SUPABASE_URL
"""
import os, io, json, re, time, datetime, hashlib, secrets
from collections import defaultdict
from typing import Optional, List, Any

import httpx
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Body, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse

__version__ = "1.5.12"
BUILD = "2026-07-22-28"

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_KEY", "")
SCHEMA = "ops"

app = FastAPI(title="OPS DC Ship Forecast", version=__version__)
import logging
log = logging.getLogger("ops")
logging.basicConfig(level=logging.INFO)
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException

def _cors_json(status, payload):
    return JSONResponse(status_code=status, content=payload, headers={
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control-Allow-Methods": "*",
        "Access-Control-Expose-Headers": "*",
    })

def _db_hint(msg):
    """Turn a raw database error into the specific migration step that fixes it."""
    low = (msg or "").lower()
    if "on conflict" in low or "unique or exclusion constraint" in low or "42p10" in low:
        return ("  FIX: the version-scoped unique indexes are missing. Run the v1.5.10 block at the "
                "end of ops_schema.sql, then ops_grants.sql, then:  NOTIFY pgrst, 'reload schema';")
    if "does not exist" in low and "column" in low:
        return ("  FIX: the database is missing columns added in v1.5.6. Run ops_schema.sql, then "
                "ops_grants.sql, then:  NOTIFY pgrst, 'reload schema';")
    if "duplicate key" in low or "23505" in low:
        return ("  FIX: an obsolete cycle-scoped unique constraint is still present. Run the v1.5.10 "
                "block at the end of ops_schema.sql, then:  NOTIFY pgrst, 'reload schema';")
    if "permission denied" in low or "42501" in low:
        return "  FIX: run ops_grants.sql, then:  NOTIFY pgrst, 'reload schema';"
    return ""

@app.exception_handler(StarletteHTTPException)
async def _http_exc_handler(request, exc):
    d = exc.detail
    if isinstance(d, str) and exc.status_code >= 500:
        d = d + _db_hint(d)
    return _cors_json(exc.status_code, {"detail": d})

@app.exception_handler(RequestValidationError)
async def _validation_exc_handler(request, exc):
    return _cors_json(422, {"detail": f"Invalid request: {exc.errors()[:3]}"})

@app.exception_handler(Exception)
async def _unhandled_exc_handler(request, exc):
    import traceback
    tb = traceback.format_exc()
    log.error("UNHANDLED %s %s\n%s", request.method, request.url.path, tb)
    msg = f"{type(exc).__name__}: {exc}"
    return _cors_json(500, {"detail": msg[:600] + _db_hint(msg)})

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_origin_regex=".*",
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# ------------------------------------------------------------------
# DB layer — PostgREST client with in-memory fallback (demo/test mode)
# ------------------------------------------------------------------
def _hash_pw(pw, salt=None):
    salt = salt or secrets.token_hex(8)
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 100000).hex()
    return f"{salt}${h}"

def _verify_pw(pw, stored):
    try:
        salt, h = stored.split("$", 1)
        return secrets.compare_digest(hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 100000).hex(), h)
    except Exception:
        return False

class MemoryDB:
    def __init__(self):
        self.t = defaultdict(list)
        self.seq = defaultdict(int)
        self._seed()

    def _seed(self):
        for i, (c, n, k, s) in enumerate([("ATL","Atlanta","domestic",1),("MX","Mexico","domestic",2),
                                          ("IT","Italy","export",3),("CN","China","export",4),("TH","Thailand","export",5)],1):
            self.t["dc"].append({"id":i,"code":c,"name":n,"kind":k,"sort":s,"active":True})
        for p,d in [("US","ATL"),("CA","ATL"),("MX","MX"),("IT","IT"),("CN","CN"),("TH","TH")]:
            self.insert("plant_alias", {"prefix":p,"dc_code":d})
        for i,(c,n) in enumerate([("IND","Independents"),("KA","Key Account"),("ICP","ICP"),("DS","Dept. Stores"),
                                  ("ECOM","E-Commerce"),("BTQ","Boutique"),("RSPT","Reg Sport"),("NSPT","Nat Sport"),
                                  ("DEF","Defense"),("GOG","Total Goggles")],1):
            self.t["channel"].append({"id":i,"code":c,"name":n,"sort":i,"active":True})
        for cot,ch in [("IND US","IND"),("IND CA","IND"),("KA National","KA"),("KA CA","KA"),("KAR Optical","KA"),
                       ("KAR Sun","KA"),("ICP","ICP"),("Dept. Stores","DS"),("Dept. Stores Off Price","DS"),
                       ("Dept. Stores EOL","DS"),("Ecomm","ECOM"),("Boutiques","BTQ"),("Regional Sport","RSPT"),
                       ("National Sport","NSPT"),("Defense","DEF"),("Audiology","ECOM")]:
            self.insert("cot_alias", {"cot":cot,"channel_code":ch})
        self.insert("app_user", {"username":"admin","pw_hash":_hash_pw("ChangeMe#2026"),
                                 "role":"admin","must_change":True,"active":True})

    def select(self, table, filters=None, order=None):
        rows = self.t[table]
        for k, v in (filters or {}).items():
            rows = [r for r in rows if r.get(k) == v]
        if order:
            desc = order.startswith("-")
            key = order.lstrip("-")
            rows = sorted(rows, key=lambda r: (r.get(key) is None, r.get(key)), reverse=desc)
        return [dict(r) for r in rows]

    def insert(self, table, row):
        self.seq[table] += 1
        r = dict(row); r.setdefault("id", self.seq[table])
        self.seq[table] = max(self.seq[table], r["id"])
        self.t[table].append(r)
        return dict(r)

    def upsert(self, table, rows, keys):
        out = []
        for row in rows:
            match = None
            for r in self.t[table]:
                if all(r.get(k) == row.get(k) for k in keys):
                    match = r; break
            if match:
                match.update(row); out.append(dict(match))
            else:
                out.append(self.insert(table, row))
        return out

    def _raise(self, r, table, op):
        """PostgREST puts the real Postgres error in the response body; raise_for_status()
        throws it away. Surface it so a failure explains itself."""
        try:
            body = r.json()
            detail = "; ".join(str(body.get(k)) for k in ("message", "details", "hint", "code")
                               if body.get(k))
        except Exception:
            detail = (r.text or "")[:300]
        raise HTTPException(502, f"Database rejected {op} on '{table}' (HTTP {r.status_code}): {detail}")

    def has_column(self, table, col):
        return True   # in-memory store is schemaless

    def update(self, table, filters, patch):
        n = 0
        for r in self.t[table]:
            if all(r.get(k) == v for k, v in filters.items()):
                r.update(patch); n += 1
        return n

    def delete(self, table, filters):
        before = len(self.t[table])
        self.t[table] = [r for r in self.t[table] if not all(r.get(k) == v for k, v in filters.items())]
        return before - len(self.t[table])


class RestDB:
    def __init__(self):
        self.base = f"{SUPABASE_URL}/rest/v1"
        self.h = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
                  "Accept-Profile": SCHEMA, "Content-Profile": SCHEMA}
        self.client = httpx.Client(timeout=60)

    def _params(self, filters):
        return {k: f"eq.{v}" for k, v in (filters or {}).items()}

    def has_column(self, table, col):
        """Ask PostgREST for a single row projecting only `col`. A missing column (or a stale
        PostgREST schema cache after a migration) makes this fail, which is exactly what we
        want to detect before an import blows up mid-write."""
        try:
            r = self.client.get(f"{self.base}/{table}", headers=self.h,
                                params={"select": col, "limit": 1})
            return r.status_code < 400
        except Exception:
            return False

    def select(self, table, filters=None, order=None):
        params = self._params(filters)
        if order:
            params["order"] = order.lstrip("-") + (".desc" if order.startswith("-") else ".asc")
        r = self.client.get(f"{self.base}/{table}", headers=self.h, params=params)
        if r.status_code >= 400: self._raise(r, table, "select")
        return r.json()

    def insert(self, table, row):
        h = dict(self.h); h["Prefer"] = "return=representation"
        r = self.client.post(f"{self.base}/{table}", headers=h, json=row)
        if r.status_code >= 400: self._raise(r, table, "insert")
        return r.json()[0]

    def upsert(self, table, rows, keys):
        if not rows: return []
        h = dict(self.h); h["Prefer"] = "return=representation,resolution=merge-duplicates"
        params = {"on_conflict": ",".join(keys)}
        out = []
        for i in range(0, len(rows), 500):  # homogeneous batches, chunked
            r = self.client.post(f"{self.base}/{table}", headers=h, params=params, json=rows[i:i+500])
            if r.status_code >= 400: self._raise(r, table, "upsert")
            out.extend(r.json())
        return out

    def update(self, table, filters, patch):
        r = self.client.patch(f"{self.base}/{table}", headers=self.h, params=self._params(filters), json=patch)
        r.raise_for_status(); return 1

    def delete(self, table, filters):
        r = self.client.delete(f"{self.base}/{table}", headers=self.h, params=self._params(filters))
        r.raise_for_status(); return 1


DB = RestDB() if SUPABASE_URL else MemoryDB()
DB_MODE = "supabase" if SUPABASE_URL else "memory"

def ensure_seed_admin():
    try:
        rows = DB.select("app_user", {"username": "admin"})
        if not rows:
            DB.insert("app_user", {"username": "admin", "pw_hash": _hash_pw("ChangeMe#2026"),
                                   "role": "admin", "must_change": True, "active": True,
                                   "created_at": datetime.datetime.utcnow().isoformat()})
    except Exception as e:
        print("seed admin skipped:", e)
ensure_seed_admin()

def _new_token(): return secrets.token_urlsafe(24)

def require_role(authorization, allowed):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u: raise HTTPException(401, "Not authenticated")
    if u["role"] not in allowed:
        raise HTTPException(403, f"Your role ({u['role']}) cannot perform this action")
    return u

def current_user(token):
    if not token: return None
    s = DB.select("session", {"token": token})
    if not s: return None
    s = s[0]
    try:
        if datetime.datetime.fromisoformat(s["expires_at"].replace("Z","")) < datetime.datetime.utcnow():
            return None
    except Exception:
        pass
    u = DB.select("app_user", {"id": s["user_id"]})
    return u[0] if u else None

# special (non-channel) forecast lines
LINES = ["RW_TTL", "RW_META", "NUANCE", "OW_TTL", "OW_META"]
WEARABLE_BRANDS = {"RW": "RW", "OW": "OW", "AW": "NUANCE", "YM": "NUANCE"}
GOGGLE_MIX = {"SNOW  GOGGLES", "SNOW GOGGLES", "MX GOGGLE"}

# NA column map — stored as setting so future layout drift is config, not code
NA_MAP_DEFAULT = {
    "sheet": "2026 FC", "week_col": "A",
    "channels": {
        "IND": ["CC:CJ"], "KA": ["CR:CU"], "ICP": ["BJ:BL"],
        "DS": ["R:T", "W:Y"], "ECOM": ["U", "Z"], "BTQ": ["V", "AA"],
        "RSPT": ["AB:AK"], "NSPT": ["AL:AU"], "DEF": ["AY"], "GOG": ["AV:AX"],
    },
    "rw_sheet": "2026 RW FC", "rw_total": "S", "rw_meta": "L",
    "aw_sheet": "2026 AW FC", "nuance_total": "P",
    "ow_sheet": "2026 OW FC", "ow_total": "T", "ow_meta": "L",
}

def get_setting(key, default=None):
    rows = DB.select("app_setting", {"key": key})
    return rows[0]["value"] if rows else default

def set_setting(key, value):
    DB.upsert("app_setting", [{"key": key, "value": value}], ["key"])

# ------------------------------------------------------------------
# helpers
# ------------------------------------------------------------------
def col_letters_to_idx(letter):
    n = 0
    for ch in letter.strip().upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1

def expand_ranges(specs):
    idxs = []
    for spec in specs:
        if ":" in spec:
            a, b = spec.split(":")
            idxs.extend(range(col_letters_to_idx(a), col_letters_to_idx(b) + 1))
        else:
            idxs.append(col_letters_to_idx(spec))
    return idxs

def parse_period_key(val):
    """'7-26' -> (7,26). Accepts int-ish strings."""
    if val is None: return None
    m = re.match(r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*$", str(val))
    if not m: return None
    return int(m.group(1)), int(m.group(2))

def month_to_quarter(m): return (int(m) - 1) // 3 + 1

def ensure_cycle(year, quarter):
    rows = DB.select("cycle", {"year": year, "quarter": quarter})
    if rows: return rows[0]
    return DB.insert("cycle", {"year": year, "quarter": quarter, "label": f"{year} Q{quarter}"})

def ensure_period(cycle_id, month_no, week_no):
    rows = DB.select("period", {"cycle_id": cycle_id, "month_no": month_no, "week_no": week_no})
    if rows: return rows[0]
    return DB.insert("period", {"cycle_id": cycle_id, "month_no": month_no, "week_no": week_no,
                                "label": f"{month_no}-{week_no}", "sort": month_no * 100 + week_no})

def ensure_periods_bulk(cycle_id, pairs):
    """Resolve many (month_no, week_no) pairs to period rows with ONE select and ONE batched
    insert. Replaces per-cell ensure_period() calls, which cost an HTTP round trip each against
    a hosted Postgres and were the dominant import cost. Returns {(mo, wk): period_row}."""
    pairs = {(int(m), int(w)) for m, w in pairs}
    existing = DB.select("period", {"cycle_id": cycle_id})
    have = {(r["month_no"], r["week_no"]): r for r in existing}
    missing = [p for p in pairs if p not in have]
    if missing:
        new_rows = [{"cycle_id": cycle_id, "month_no": m, "week_no": w,
                     "label": f"{m}-{w}", "sort": m * 100 + w} for (m, w) in sorted(missing)]
        DB.upsert("period", new_rows, ["cycle_id", "month_no", "week_no"])
        for r in DB.select("period", {"cycle_id": cycle_id}):
            have[(r["month_no"], r["week_no"])] = r
    return {p: have[p] for p in pairs if p in have}

def plant_to_dc(plant, aliases, _cache=None):
    if not plant: return None
    p = str(plant).upper()
    if _cache is not None and p in _cache: return _cache[p]
    best = None
    for a in aliases:
        if p.startswith(a["prefix"]) and (best is None or len(a["prefix"]) > len(best["prefix"])):
            best = a
    res = best["dc_code"] if best else None
    if _cache is not None: _cache[p] = res
    return res

def load_wb(data):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

def num(v):
    try:
        if v is None or v == "": return 0.0
        return float(v)
    except Exception:
        return 0.0

# ------------------------------------------------------------------
# meta endpoints
# ------------------------------------------------------------------
@app.get("/version")
def version():
    return {"version": __version__, "build": BUILD, "db_mode": DB_MODE, "schema": SCHEMA}

@app.get("/health")
def health():
    return {"ok": True, "time": datetime.datetime.utcnow().isoformat()}

@app.get("/bootstrap")
def bootstrap():
    return {
        "dcs": DB.select("dc", order="sort"),
        "channels": DB.select("channel", order="sort"),
        "customers": DB.select("customer", order="sort"),
        "cycles": DB.select("cycle", order="-id"),
        "versions": DB.select("version", order="-id"),
        "plant_aliases": DB.select("plant_alias"),
        "cot_aliases": DB.select("cot_alias"),
        "na_map": get_setting("na_map", NA_MAP_DEFAULT),
        "app": {"version": __version__, "build": BUILD, "db_mode": DB_MODE},
    }

# ------------------------------------------------------------------
# admin CRUD (thin, id-based)
# ------------------------------------------------------------------
@app.post("/admin/customer")
def add_customer(payload: dict = Body(...)):
    name = (payload.get("name") or "").strip()
    if not name: raise HTTPException(400, "name required")
    existing = DB.select("customer", {"name": name})
    if existing: return existing[0]
    sorts = [c.get("sort", 0) for c in DB.select("customer")] or [0]
    return DB.insert("customer", {"name": name, "aliases": payload.get("aliases", []),
                                  "sort": max(sorts) + 1, "active": True})

@app.patch("/admin/customer/{cid}")
def patch_customer(cid: int, payload: dict = Body(...)):
    allowed = {k: v for k, v in payload.items() if k in ("name", "active", "sort", "aliases")}
    DB.update("customer", {"id": cid}, allowed)
    return {"ok": True}

@app.post("/admin/remove_customer")
def admin_remove_customer(payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    """Deactivate customer(s) by name (admin only). Additive/safe: sets active=False, never deletes rows.
    Used to clean up stray entities like a country name mistakenly imported as a customer."""
    require_role(authorization, ("admin",))
    names = payload.get("names") or ([payload["name"]] if payload.get("name") else [])
    targets = {str(n).strip().lower() for n in names}
    removed = []
    for c in DB.select("customer"):
        if c["name"].strip().lower() in targets and c.get("active", True):
            DB.update("customer", {"id": c["id"]}, {"active": False})
            removed.append(c["name"])
    return {"ok": True, "removed": removed}

@app.get("/admin/stray_customers")
def admin_stray_customers(authorization: Optional[str] = Header(None)):
    """List customers whose names look like country/total headers (not real customers)."""
    require_role(authorization, ("admin",))
    BAD = {"italy", "china", "thailand", "direct", "total", "atlanta", "mexico",
           "total italy", "total china", "total thailand", "total direct", "grand total", "sum"}
    strays = [{"id": c["id"], "name": c["name"]} for c in DB.select("customer")
              if c.get("active", True) and c["name"].strip().lower() in BAD]
    return {"strays": strays}

@app.post("/admin/cot_alias")
def add_cot(payload: dict = Body(...)):
    return DB.upsert("cot_alias", [{"cot": payload["cot"], "channel_code": payload["channel_code"]}], ["cot"])[0]

@app.post("/admin/plant_alias")
def add_plant(payload: dict = Body(...)):
    return DB.upsert("plant_alias", [{"prefix": payload["prefix"].upper(), "dc_code": payload["dc_code"]}], ["prefix"])[0]

@app.post("/admin/na_map")
def save_na_map(payload: dict = Body(...)):
    set_setting("na_map", payload); return {"ok": True}

# ------------------------------------------------------------------
# versions
# ------------------------------------------------------------------
@app.post("/version/create")
def create_version(payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    require_role(authorization, ("admin","planner"))
    year, quarter = int(payload["year"]), int(payload["quarter"])
    cyc = ensure_cycle(year, quarter)
    tag = payload.get("week_tag") or f"WK{payload.get('week', '')}"
    v = DB.insert("version", {"cycle_id": cyc["id"], "week_tag": tag, "kind": payload.get("kind", "WEEKLY"),
                              "note": payload.get("note", ""), "published": False,
                              "created_at": datetime.datetime.utcnow().isoformat()})
    src = payload.get("copy_from")
    if src:
        for table, keys in [("penetration", ["version_id", "dc_code", "target_kind", "target_key"]),
                            ("rate", ["version_id", "dc_code", "kind"]),
                            ("dummy_plan", ["version_id", "period_id", "dc_code"]),
                            ("forecast_channel", ["version_id", "period_id", "channel_code"]),
                            ("forecast_customer", ["version_id", "period_id", "customer_id"])]:
            rows = DB.select(table, {"version_id": int(src)})
            cp = []
            for r in rows:
                r = dict(r); r.pop("id", None); r["version_id"] = v["id"]
                cp.append(r)
            if cp: DB.upsert(table, cp, keys)
    return v

@app.post("/version/{vid}/publish")
def publish_version(vid: int, authorization: Optional[str] = Header(None)):
    require_role(authorization, ("admin","planner"))
    DB.update("version", {"id": vid}, {"published": True, "published_at": datetime.datetime.utcnow().isoformat()})
    return {"ok": True}

@app.delete("/version/{vid}")
def delete_version(vid: int, authorization: Optional[str] = Header(None)):
    """Delete a single WEEK (version) and EVERYTHING imported for it — each week is a full,
    independent snapshot: forecast, customers, penetration, rates, dummy plan, overrides, AND
    this week's own actuals/accessories/dummies. Other weeks are untouched. Protected accounts
    are never involved here."""
    require_role(authorization, ("admin", "planner"))
    v = DB.select("version", {"id": vid})
    if not v: raise HTTPException(404, "version not found")
    for table in ("forecast_channel", "forecast_customer", "penetration", "rate",
                  "dummy_plan", "override",
                  "actual", "actual_customer", "actual_dim_country", "actual_cust_dim"):
        DB.delete(table, {"version_id": vid})
    DB.delete("version", {"id": vid})
    return {"ok": True, "deleted": "version", "week_tag": v[0].get("week_tag")}

@app.delete("/cycle/{cid}")
def delete_cycle(cid: int, authorization: Optional[str] = Header(None)):
    """Delete an ENTIRE CYCLE (quarter): every week/version in it AND the cycle-level uploads
    (actuals, accessories, dummies, customer x DC penetration data, periods). This wipes
    everything uploaded for that quarter. Admin only."""
    require_role(authorization, ("admin",))
    cyc = DB.select("cycle", {"id": cid})
    if not cyc: raise HTTPException(404, "cycle not found")
    # delete all versions in this cycle (and their per-week data incl. actuals)
    for v in DB.select("version", {"cycle_id": cid}):
        for table in ("forecast_channel", "forecast_customer", "penetration", "rate",
                      "dummy_plan", "override",
                      "actual", "actual_customer", "actual_dim_country", "actual_cust_dim"):
            DB.delete(table, {"version_id": v["id"]})
        DB.delete("version", {"id": v["id"]})
    # sweep any legacy cycle-scoped actuals (pre-1.5.6 rows without version_id) + periods
    for table in ("actual", "actual_customer", "actual_dim_country", "actual_cust_dim", "period"):
        DB.delete(table, {"cycle_id": cid})
    DB.delete("cycle", {"id": cid})
    return {"ok": True, "deleted": "cycle", "label": cyc[0].get("label")}

@app.get("/version/{vid}/periods")
def version_periods(vid: int):
    v = DB.select("version", {"id": vid})
    if not v: raise HTTPException(404, "version not found")
    return sorted(DB.select("period", {"cycle_id": v[0]["cycle_id"]}), key=lambda p: p["sort"])

# ------------------------------------------------------------------
# imports
# ------------------------------------------------------------------
def _clear_import(kind, version_id=None, cycle_id=None):
    """Clean-replace: remove existing rows for a given import type before re-committing.
    Scoped precisely so other imports/actuals are never touched."""
    if kind == "NA" and version_id is not None:
        DB.delete("forecast_channel", {"version_id": version_id})
    elif kind == "CUSTFC" and version_id is not None:
        DB.delete("forecast_customer", {"version_id": version_id})
    elif kind in ("ACTUALS", "ACC", "DUMMY") and version_id is not None:
        # actual table shares one table across kinds — delete only the matching kind(s), THIS week only
        if kind == "ACTUALS":
            for k in ("GROSS", "RETURN"):
                DB.delete("actual", {"version_id": version_id, "kind": k})
        elif kind == "ACC":
            DB.delete("actual", {"version_id": version_id, "kind": "ACC"})
        elif kind == "DUMMY":
            DB.delete("actual", {"version_id": version_id, "kind": "DUMMY"})

def _already_imported(kind, version_id=None, cycle_id=None):
    """Has this import type already been committed for this cycle/version?"""
    try:
        if kind == "NA" and version_id is not None:
            return len(DB.select("forecast_channel", {"version_id": version_id})) > 0
        if kind == "CUSTFC" and version_id is not None:
            return len(DB.select("forecast_customer", {"version_id": version_id})) > 0
        if kind == "ACTUALS" and version_id is not None:
            return any(a["kind"] in ("GROSS", "RETURN") for a in DB.select("actual", {"version_id": version_id}))
        if kind == "ACC" and version_id is not None:
            return any(a["kind"] == "ACC" for a in DB.select("actual", {"version_id": version_id}))
        if kind == "DUMMY" and version_id is not None:
            return any(a["kind"] == "DUMMY" for a in DB.select("actual", {"version_id": version_id}))
    except Exception:
        pass
    return False

def _commit_or_preview(commit, kind, filename, rows_payload, do_commit):
    if not commit:
        return {"preview": True, **rows_payload}
    result = do_commit()
    DB.insert("import_log", {"kind": kind, "filename": filename, "rows": result.get("rows", 0),
                             "meta": result.get("meta", {}), "created_at": datetime.datetime.utcnow().isoformat()})
    return {"preview": False, **result}

REQUIRED_COLUMNS = [
    ("actual", "version_id"),
    ("actual_customer", "version_id"),
    ("actual_dim_country", "version_id"),
    ("actual_cust_dim", "version_id"),
    ("customer_alias", "actuals_name"),
]
_SCHEMA_CACHE = {"missing": None, "ts": 0.0}

def schema_missing(force=False):
    """List required table.column pairs the live database is missing. Cached for 60s so import
    pre-flight checks stay cheap."""
    now = time.time()
    if not force and _SCHEMA_CACHE["missing"] is not None and now - _SCHEMA_CACHE["ts"] < 60:
        return _SCHEMA_CACHE["missing"]
    miss = []
    for t, c in REQUIRED_COLUMNS:
        try:
            if not DB.has_column(t, c): miss.append(f"{t}.{c}")
        except Exception:
            miss.append(f"{t}.{c}")
    _SCHEMA_CACHE["missing"] = miss; _SCHEMA_CACHE["ts"] = now
    return miss

def require_schema():
    """Fail an import early with an actionable message instead of a cryptic mid-write 500."""
    miss = schema_missing()
    if miss:
        raise HTTPException(400,
            "Database is not migrated for v1.5.6+. Missing: " + ", ".join(miss) +
            ". Fix: run ops_schema.sql, then ops_grants.sql, then reload the PostgREST schema "
            "cache with:  NOTIFY pgrst, 'reload schema';  (or restart the PostgREST service). "
            "Then retry this import.")

@app.get("/health/schema")
def health_schema(force: bool = False):
    miss = schema_missing(force=force)
    return {"ok": not miss, "missing": miss, "db_mode": DB_MODE, "schema": SCHEMA,
            "required": [f"{t}.{c}" for t, c in REQUIRED_COLUMNS],
            "fix": None if not miss else
                   "Run ops_schema.sql, then ops_grants.sql, then: NOTIFY pgrst, 'reload schema';"}

@app.get("/health/writecheck")
def health_writecheck(version_id: int, authorization: Optional[str] = Header(None)):
    """Actually perform the upsert that imports rely on, using a sentinel row, then remove it.
    A column check alone cannot catch a missing unique index — this does, before you spend
    minutes uploading a large file."""
    require_role(authorization, ("admin", "planner"))
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    cyc = DB.select("cycle", {"id": v[0]["cycle_id"]})[0]
    pers = sorted(DB.select("period", {"cycle_id": cyc["id"]}), key=lambda p: p.get("sort", 0))
    if not pers:
        pers = [ensure_periods_bulk(cyc["id"], {(1, 1)})[(1, 1)]]
    pid = pers[0]["id"]
    SENTINEL = "__WRITECHECK__"
    results = {}
    try:
        DB.upsert("actual", [{"version_id": version_id, "cycle_id": cyc["id"], "period_id": pid,
                              "dc_code": "ATL", "channel_code": SENTINEL, "kind": SENTINEL,
                              "units": 0}],
                  ["version_id", "period_id", "dc_code", "channel_code", "kind"])
        # write twice: proves the conflict target resolves instead of duplicating
        DB.upsert("actual", [{"version_id": version_id, "cycle_id": cyc["id"], "period_id": pid,
                              "dc_code": "ATL", "channel_code": SENTINEL, "kind": SENTINEL,
                              "units": 0}],
                  ["version_id", "period_id", "dc_code", "channel_code", "kind"])
        results["actual"] = "ok"
    except HTTPException as e:
        results["actual"] = f"FAILED: {e.detail}"
    except Exception as e:
        results["actual"] = f"FAILED: {type(e).__name__}: {e}"
    finally:
        try: DB.delete("actual", {"version_id": version_id, "kind": SENTINEL})
        except Exception: pass
    ok = all(str(x) == "ok" for x in results.values())
    return {"ok": ok, "checks": results, "missing_columns": schema_missing(force=True),
            "fix": None if ok else
                   "Run ops_schema.sql (including the v1.5.10 block at the end), then "
                   "ops_grants.sql, then:  NOTIFY pgrst, 'reload schema';"}

@app.get("/version/{vid}/readiness")
def version_readiness(vid: int):
    """What this week actually has, and what is still required for the forecast to be non-zero.
    Each week is an independent snapshot, so every input must be present for THIS week."""
    v = DB.select("version", {"id": vid})
    if not v: raise HTTPException(404, "version not found")
    fc = DB.select("forecast_channel", {"version_id": vid})
    fcu = DB.select("forecast_customer", {"version_id": vid})
    acts = DB.select("actual", {"version_id": vid})
    pen = DB.select("penetration", {"version_id": vid})
    rates = DB.select("rate", {"version_id": vid})
    dpl = DB.select("dummy_plan", {"version_id": vid})
    pen_ch = [p for p in pen if p["target_kind"] == "CHANNEL" and (p.get("pct") or 0) > 0]
    pen_cu = [p for p in pen if p["target_kind"] == "CUSTOMER" and (p.get("pct") or 0) > 0]
    dom_pen = [p for p in pen_ch if p["dc_code"] in ("ATL", "MX")]
    items = [
        {"key": "na", "label": "NA planner forecast", "ok": len(fc) > 0, "count": len(fc),
         "why": "Without it every forecast week is 0.", "fix": "Imports → 1 · NA file → Commit"},
        {"key": "actuals", "label": "Actuals (All Channels)", "ok": any(a["kind"] in ("GROSS", "RETURN") for a in acts),
         "count": len(acts), "why": "Sets which weeks are actualized and feeds derived penetration.",
         "fix": "Imports → 2 · Actuals → Commit"},
        {"key": "pen_channel", "label": "Penetration % SAVED for Atlanta/Mexico", "ok": len(dom_pen) > 0,
         "count": len(dom_pen),
         "why": "Forecast week = penetration % x NA forecast. With none saved, Atlanta and Mexico forecast weeks are 0 and everything falls into Direct.",
         "fix": "Penetration % → Suggest from actuals → then click SAVE ALL %"},
        {"key": "customer_list", "label": "Customer list defined", "ok": len(DB.select("customer")) > 0,
         "count": len(DB.select("customer")),
         "why": "Customer columns come from this list; names are matched to the BO actuals.",
         "fix": "Admin → Customers (or import a Direct file once to seed the names)"},
        {"key": "customers", "label": "Customer forecast (Direct file) — OPTIONAL",
         "ok": True, "count": len(fcu),
         "why": ("Not required. With no file, customer volumes are derived from this week's BO actuals "
                 "(each customer's share of export volume applied to the Direct residual). Import a file "
                 "only if you want to override that with planner numbers."),
         "fix": "Optional: Imports → 5 · Customer forecast → Commit"},
        {"key": "pen_customer", "label": "Penetration % SAVED for customers — OPTIONAL",
         "ok": True, "count": len(pen_cu),
         "why": ("Not required. Without it the country split falls back to each customer's own actual "
                 "IT/CN/TH mix. Save penetration to override with the full brand/mix/channel ladder."),
         "fix": "Optional: Penetration % → Suggest from actuals → Save all %"},
        {"key": "rates", "label": "Rates (returns/accessories)", "ok": len(rates) > 0, "count": len(rates),
         "why": "Optional, but returns and accessories forecasts stay 0 without them.",
         "fix": "Rates & dummies → Save"},
        {"key": "dummies", "label": "Dummy plan", "ok": len(dpl) > 0, "count": len(dpl),
         "why": "Optional; dummy units per week.", "fix": "Rates & dummies → Save"},
    ]
    blocking = [i for i in items if not i["ok"] and i["key"] in
                ("na", "actuals", "pen_channel", "customer_list")]
    return {"version_id": vid, "week_tag": v[0].get("week_tag"), "items": items,
            "ok": not blocking, "blocking": [i["key"] for i in blocking]}

@app.get("/import/status")
def import_status(version_id: int, kind: str):
    """Lightweight check for whether an import type is already committed for this week.
    Replaces the old 'peek' that re-uploaded and re-parsed the whole file just to read this
    flag — that doubled upload time and parse time on every commit."""
    k = (kind or "").upper()
    if k not in ("NA", "CUSTFC", "ACTUALS", "ACC", "DUMMY"):
        raise HTTPException(400, "unknown import kind")
    # Actuals / accessories / dummies never warn on re-import (standing rule: a fresh week is a
    # true clean slate and re-committing simply clean-replaces). Mirrors the hardcoded
    # already_imported=False in those importers' previews, so behaviour is unchanged.
    if k in ("ACTUALS", "ACC", "DUMMY"):
        return {"kind": k, "already_imported": False}
    return {"kind": k, "already_imported": _already_imported(k, version_id=version_id)}

@app.post("/import/na")
async def import_na(file: UploadFile = File(...), version_id: int = Form(...), commit: bool = Form(False), authorization: Optional[str] = Header(None)):
    require_role(authorization, ('admin','planner'))
    data = await file.read()
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    v = v[0]
    cyc = DB.select("cycle", {"id": v["cycle_id"]})[0]
    q_months = {(cyc["quarter"] - 1) * 3 + i for i in (1, 2, 3)}
    m = get_setting("na_map", NA_MAP_DEFAULT)
    wb = load_wb(data)
    missing = [s for s in [m["sheet"], m["rw_sheet"], m["aw_sheet"], m["ow_sheet"]] if s not in wb.sheetnames]
    if missing: raise HTTPException(400, f"NA file missing sheets: {missing}")

    ch_idx = {ch: expand_ranges(specs) for ch, specs in m["channels"].items()}
    grid = defaultdict(dict)   # pkey -> {code: units}
    ws = wb[m["sheet"]]
    wkc = col_letters_to_idx(m["week_col"])
    for row in ws.iter_rows(min_row=4, values_only=True):
        pk = parse_period_key(row[wkc] if wkc < len(row) else None)
        if not pk or pk[0] not in q_months: continue
        for ch, idxs in ch_idx.items():
            grid[pk][ch] = grid[pk].get(ch, 0) + sum(num(row[i]) for i in idxs if i < len(row))
    def line_sheet(sheet, total_col, meta_col, tkey, mkey):
        ws2 = wb[sheet]
        ti, mi = col_letters_to_idx(total_col), (col_letters_to_idx(meta_col) if meta_col else None)
        for row in ws2.iter_rows(min_row=4, values_only=True):
            pk = parse_period_key(row[0])
            if not pk or pk[0] not in q_months: continue
            grid[pk][tkey] = num(row[ti]) if ti < len(row) else 0
            if mi is not None:
                grid[pk][mkey] = num(row[mi]) if mi < len(row) else 0
    line_sheet(m["rw_sheet"], m["rw_total"], m["rw_meta"], "RW_TTL", "RW_META")
    line_sheet(m["aw_sheet"], m["nuance_total"], None, "NUANCE", "")
    line_sheet(m["ow_sheet"], m["ow_total"], m["ow_meta"], "OW_TTL", "OW_META")

    ch_codes = [c["code"] for c in DB.select("channel")]
    totals = {c: round(sum(g.get(c, 0) for g in grid.values())) for c in ch_codes + LINES}
    preview = {"periods": len(grid), "totals": totals,
               "grand_total": round(sum(totals[c] for c in ch_codes)),
               "already_imported": _already_imported("NA", version_id=version_id)}

    def do_commit():
        _clear_import("NA", version_id=version_id)   # clean replace
        rows = []
        _pmap = ensure_periods_bulk(cyc["id"], grid.keys())
        for (mo, wk), vals in grid.items():
            per = _pmap[(mo, wk)]
            for code, units in vals.items():
                if code == "": continue
                rows.append({"version_id": version_id, "period_id": per["id"], "channel_code": code, "units": units})
        DB.upsert("forecast_channel", rows, ["version_id", "period_id", "channel_code"])
        return {"rows": len(rows), "meta": preview}
    return _commit_or_preview(commit, "NA", file.filename, preview, do_commit)


def parse_all_channels(data, filename, aliases, cot_map, month_filter=None, want_entity=False, want_dims=False):
    """Shared parser for BO 'All Channels Data' exports.
    Returns (agg, cust_agg, unmapped_cot, unmapped_plants, nrows) where
    agg[(year,mo,wk,dc,ch,kind)] = units and cust_agg[(year,mo,wk,dc,entity_upper)] = units."""
    def rows_iter():
        if filename.lower().endswith(".csv"):
            import csv
            text = data.decode("utf-8-sig", errors="replace").splitlines()
            yield from csv.reader(text)
        else:
            wb = load_wb(data)
            name = "All Channels Data" if "All Channels Data" in wb.sheetnames else wb.sheetnames[0]
            for r in wb[name].iter_rows(values_only=True):
                yield r
    it = rows_iter()
    header = [str(h or "").strip() for h in next(it)]
    def find(*names):
        for n in names:
            for i, h in enumerate(header):
                if h.lower() == n.lower(): return i
        return None
    iY, iM, iW = find("Year"), find("Month"), find("Week")
    iInv = find("Invoices +", "Invoices")
    iPl = find("DI: Original Plant Code", "Plant")
    iTK = find("Transaction Kind")
    iCOT = find("COT")
    iBr = find("Brand: Code")
    iMix = find("Collection Mix: Desc", "Collection Mix")
    iEnt = find("Commercial Entity: Name", "Customer: Name", "Customer")
    need = {"Year": iY, "Month": iM, "Week": iW, "Invoices": iInv, "Plant": iPl,
            "Transaction Kind": iTK, "COT": iCOT}
    missing = [k for k, v2 in need.items() if v2 is None]
    if missing: raise HTTPException(400, f"File missing columns: {missing}")
    agg, cust_agg = defaultdict(float), defaultdict(float)
    dim_country = defaultdict(float)   # (dimkind, dimval, dc) -> units  [export dc only]
    cust_dim = defaultdict(float)      # (cust_upper, dimkind, dimval) -> units  [ALL dcs, for mix weighting]
    unmapped_cot, unmapped_plant = set(), set()
    nrows = 0
    _empty_run = 0
    _dc_cache = {}
    for r in it:
        # BO/Excel exports pad the sheet with ~1M phantom empty rows; stop after a long empty run
        # so we don't scan the whole padded range (data is contiguous).
        if r is None or len(r) <= iInv or r[iY] in (None, ""):
            _empty_run += 1
            if _empty_run >= 1000: break
            continue
        _empty_run = 0
        try: yr, mo, wk = int(r[iY]), int(r[iM]), int(r[iW])
        except Exception: continue
        if month_filter and mo not in month_filter: continue
        units = num(r[iInv])
        if units == 0: continue
        dc = plant_to_dc(r[iPl], aliases, _dc_cache)
        if not dc:
            unmapped_plant.add(str(r[iPl])); continue
        tk = str(r[iTK] or "").upper()
        brand = str(r[iBr] or "").upper() if iBr is not None else ""
        cot = str(r[iCOT] or "")
        mix = str(r[iMix] or "").upper() if iMix is not None else ""
        nrows += 1
        if cot == "Meta LLC" or brand in WEARABLE_BRANDS:
            line = "META_RW" if (cot == "Meta LLC" and brand == "RW") else \
                   "META_OW" if (cot == "Meta LLC" and brand == "OW") else \
                   WEARABLE_BRANDS.get(brand, "RW")
            agg[(yr, mo, wk, dc, "", line)] += units
            continue
        ch = "GOG" if (mix in GOGGLE_MIX and cot in ("Regional Sport", "National Sport")) else cot_map.get(cot)
        if not ch:
            unmapped_cot.add(cot); continue
        agg[(yr, mo, wk, dc, ch, "GROSS")] += units
        if tk == "RETURNS":
            agg[(yr, mo, wk, dc, ch, "RETURN")] += abs(units)
        if want_entity and iEnt is not None and r[iEnt]:
            cust_agg[(yr, mo, wk, dc, str(r[iEnt]).strip().upper())] += units
        if want_dims:
            ent_u = str(r[iEnt]).strip().upper() if (iEnt is not None and r[iEnt]) else ""
            if dc in ("IT", "CN", "TH"):
                if brand: dim_country[("BRAND", brand, dc)] += units
                if mix:   dim_country[("MIX", mix, dc)] += units
                if cot:   dim_country[("CHANNEL", cot, dc)] += units
            if ent_u:
                # customer's dimensional mix comes from ALL their shipments (domestic dominates)
                if brand: cust_dim[(ent_u, "BRAND", brand)] += units
                if mix:   cust_dim[(ent_u, "MIX", mix)] += units
                if cot:   cust_dim[(ent_u, "CHANNEL", cot)] += units
    if want_dims:
        return agg, cust_agg, unmapped_cot, unmapped_plant, nrows, dim_country, cust_dim
    return agg, cust_agg, unmapped_cot, unmapped_plant, nrows

@app.post("/import/actuals")
async def import_actuals(file: UploadFile = File(...), version_id: int = Form(...), commit: bool = Form(False), authorization: Optional[str] = Header(None)):
    require_role(authorization, ('admin','planner'))
    require_schema()   # fail fast with an actionable message if the DB isn't migrated
    """BO 'All Channels Data' export for the CURRENT cycle. Net invoices across all transaction kinds per COT
    (Eleanor parity), wearable brands routed to RW/OW/Nuance lines, Meta LLC to META lines, goggles carved
    out of sport COTs by collection mix. Completed weeks become gray."""
    data = await file.read()
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    cyc = DB.select("cycle", {"id": v[0]["cycle_id"]})[0]
    q_months = {(cyc["quarter"] - 1) * 3 + i for i in (1, 2, 3)}
    aliases = DB.select("plant_alias")
    cot_map = {r["cot"]: r["channel_code"] for r in DB.select("cot_alias")}
    agg, cust_agg, unmapped_cot, unmapped_plant, nrows, dim_country, cust_dim = parse_all_channels(
        data, file.filename, aliases, cot_map, month_filter=q_months, want_entity=True, want_dims=True)
    by_dc = defaultdict(float)
    for (yr, mo, wk, dc, ch, k), u in agg.items():
        if k == "GROSS": by_dc[dc] += u
    preview = {"rows_read": nrows, "cells": len(agg),
               "gross_by_dc": {k: round(v2) for k, v2 in sorted(by_dc.items())},
               "unmapped_cot": sorted(unmapped_cot), "unmapped_plants": sorted(unmapped_plant),
               "already_imported": False}
    def do_commit():
        _clear_import("ACTUALS", version_id=version_id)   # clean replace THIS week (GROSS+RETURN)
        rows = []
        _pmap = ensure_periods_bulk(cyc["id"], {(a[1], a[2]) for a in agg.keys()})
        for (yr, mo, wk, dc, ch, k), u in agg.items():
            per = _pmap[(mo, wk)]
            rows.append({"version_id": version_id, "cycle_id": cyc["id"], "period_id": per["id"], "dc_code": dc,
                         "channel_code": ch, "kind": k, "units": u})
        DB.upsert("actual", rows, ["version_id", "period_id", "dc_code", "channel_code", "kind"])
        # --- customer x DC actuals (for customer->country penetration) — per week ---
        DB.delete("actual_customer", {"version_id": version_id})   # clean replace THIS week
        crows = [{"version_id": version_id, "cycle_id": cyc["id"], "year": yr, "month_no": mo, "week_no": wk,
                  "dc_code": dc, "customer_name": ent, "units": u}
                 for (yr, mo, wk, dc, ent), u in cust_agg.items()]
        if crows:
            DB.upsert("actual_customer", crows,
                      ["version_id", "year", "month_no", "week_no", "dc_code", "customer_name"])
        # dimension -> country (export) and customer -> dimension mix (for penetration fallback) — per week
        DB.delete("actual_dim_country", {"version_id": version_id})
        DB.delete("actual_cust_dim", {"version_id": version_id})
        dc_rows = [{"version_id": version_id, "cycle_id": cyc["id"], "dim_kind": k, "dim_val": val, "dc_code": d, "units": u}
                   for (k, val, d), u in dim_country.items()]
        cd_rows = [{"version_id": version_id, "cycle_id": cyc["id"], "customer_name": ent, "dim_kind": k, "dim_val": val, "units": u}
                   for (ent, k, val), u in cust_dim.items()]
        if dc_rows: DB.upsert("actual_dim_country", dc_rows, ["version_id", "dim_kind", "dim_val", "dc_code"])
        if cd_rows: DB.upsert("actual_cust_dim", cd_rows, ["version_id", "customer_name", "dim_kind", "dim_val"])
        return {"rows": len(rows), "cust_rows": len(crows),
                "dim_rows": len(dc_rows), "custdim_rows": len(cd_rows), "meta": preview}
    return _commit_or_preview(commit, "ACTUALS", file.filename, preview, do_commit)

@app.post("/import/history")
async def import_history(files: List[UploadFile] = File(None), file: UploadFile = File(None),
                         commit: bool = Form(False), authorization: Optional[str] = Header(None)):
    require_role(authorization, ('admin','planner'))
    # accept either a single 'file' or many 'files'
    uploads = [f for f in (files or []) if f is not None]
    if file is not None: uploads.append(file)
    if not uploads: raise HTTPException(400, "No file(s) provided")
    """Historical BO 'All Channels Data' export (any months/quarters — e.g. a rolling 6-month pull).
    Feeds the trailing-window penetration suggest for channels AND customers.
    Customer rows are matched by name/alias against Commercial Entity: Name — matches are counted in the
    preview so you can confirm before committing (auto-map + confirm, never silent)."""
    aliases = DB.select("plant_alias")
    cot_map = {r["cot"]: r["channel_code"] for r in DB.select("cot_alias")}
    agg, cust_raw = defaultdict(float), defaultdict(float)
    unmapped_cot, unmapped_plant = set(), set()
    nrows = 0
    filenames = []
    for uf in uploads:
        data = await uf.read()
        filenames.append(uf.filename)
        a, c, uc, up2, nr = parse_all_channels(data, uf.filename, aliases, cot_map,
                                               month_filter=None, want_entity=True)
        for k, v in a.items(): agg[k] += v
        for k, v in c.items(): cust_raw[k] += v
        unmapped_cot |= uc; unmapped_plant |= up2; nrows += nr
    # customer matching: longest customer name/alias contained in the entity name wins
    matchers = []
    for c in DB.select("customer"):
        if not c.get("active", True): continue
        for token in [c["name"]] + list(c.get("aliases") or []):
            t = str(token).strip().upper()
            if len(t) >= 3: matchers.append((t, c["id"]))
    matchers.sort(key=lambda x: -len(x[0]))
    cust_agg = defaultdict(float); matched_entities = defaultdict(int); unmatched = defaultdict(float)
    for (yr, mo, wk, dc, ent), u in cust_raw.items():
        cid = next((cid for t, cid in matchers if t in ent), None)
        if cid:
            cust_agg[(yr, mo, wk, dc, cid)] += u; matched_entities[cid] += 1
        else:
            unmatched[ent] += u
    months_seen = sorted({(yr, mo) for (yr, mo, wk, dc, ch, k) in agg})
    top_unmatched = sorted(unmatched.items(), key=lambda x: -abs(x[1]))[:15]
    id2name = {c["id"]: c["name"] for c in DB.select("customer")}
    preview = {"rows_read": nrows, "files": filenames, "months": [f"{y}-{m:02d}" for y, m in months_seen],
               "channel_cells": len(agg), "customer_cells": len(cust_agg),
               "customer_matches": {id2name.get(cid, cid): n for cid, n in
                                    sorted(matched_entities.items(), key=lambda x: -x[1])},
               "top_unmatched_entities": [{"entity": e, "units": round(u)} for e, u in top_unmatched],
               "unmapped_cot": sorted(unmapped_cot), "unmapped_plants": sorted(unmapped_plant)}
    def do_commit():
        rows = [{"year": yr, "month_no": mo, "week_no": wk, "dc_code": dc, "channel_code": ch,
                 "kind": k, "units": u} for (yr, mo, wk, dc, ch, k), u in agg.items() if k in ("GROSS", "RETURN")]
        DB.upsert("history", rows, ["year", "month_no", "week_no", "dc_code", "channel_code", "kind"])
        crows = [{"year": yr, "month_no": mo, "week_no": wk, "dc_code": dc, "customer_id": cid, "units": u}
                 for (yr, mo, wk, dc, cid), u in cust_agg.items()]
        DB.upsert("history_customer", crows, ["year", "month_no", "week_no", "dc_code", "customer_id"])
        return {"rows": len(rows) + len(crows), "meta": preview}
    return _commit_or_preview(commit, "HISTORY", ", ".join(filenames), preview, do_commit)

@app.post("/import/accessories")
async def import_accessories(file: UploadFile = File(...), version_id: int = Form(...), commit: bool = Form(False), authorization: Optional[str] = Header(None)):
    require_role(authorization, ('admin','planner'))
    require_schema()   # fail fast with an actionable message if the DB isn't migrated
    """Flat BO accessories report: Year, Month, Week, Invoices, Plant (sheet 'Report 2' or first flat sheet)."""
    data = await file.read()
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    cyc = DB.select("cycle", {"id": v[0]["cycle_id"]})[0]
    q_months = {(cyc["quarter"] - 1) * 3 + i for i in (1, 2, 3)}
    aliases = DB.select("plant_alias")
    wb = load_wb(data)
    ws = wb["Report 2"] if "Report 2" in wb.sheetnames else wb[wb.sheetnames[-1]]
    it = ws.iter_rows(values_only=True)
    header = [str(h or "").strip().lower() for h in next(it)]
    def find(n):
        for i, h in enumerate(header):
            if n in h: return i
        return None
    iY, iM, iW, iInv, iPl = find("year"), find("month"), find("week"), find("invoice"), find("plant")
    if None in (iY, iM, iW, iInv, iPl):
        raise HTTPException(400, f"Accessories sheet needs Year/Month/Week/Invoices/Plant columns; got {header}")
    agg = defaultdict(float)
    _empty = 0; _dc_cache = {}
    for r in it:
        if r is None or r[iY] in (None, ""):
            _empty += 1
            if _empty >= 1000: break
            continue
        _empty = 0
        try: mo, wk = int(r[iM]), int(r[iW])
        except Exception: continue
        if mo not in q_months: continue
        dc = plant_to_dc(r[iPl], aliases, _dc_cache)
        if dc: agg[(mo, wk, dc)] += num(r[iInv])
    preview = {"cells": len(agg), "by_dc": {}, "already_imported": False}
    for (mo, wk, dc), u in agg.items():
        preview["by_dc"][dc] = round(preview["by_dc"].get(dc, 0) + u)
    def do_commit():
        _clear_import("ACC", version_id=version_id)
        _pmap = ensure_periods_bulk(cyc["id"], {(a[0], a[1]) for a in agg.keys()})
        rows = [{"version_id": version_id, "cycle_id": cyc["id"], "period_id": _pmap[(mo, wk)]["id"],
                 "dc_code": dc, "channel_code": "", "kind": "ACC", "units": u}
                for (mo, wk, dc), u in agg.items()]
        DB.upsert("actual", rows, ["version_id", "period_id", "dc_code", "channel_code", "kind"])
        return {"rows": len(rows), "meta": preview}
    return _commit_or_preview(commit, "ACC", file.filename, preview, do_commit)

@app.post("/import/dummies")
async def import_dummies(file: UploadFile = File(...), version_id: int = Form(...), commit: bool = Form(False), authorization: Optional[str] = Header(None)):
    require_role(authorization, ('admin','planner'))
    require_schema()   # fail fast with an actionable message if the DB isn't migrated
    """Dummy actuals from BO 'All Channels' flat sheet in the dummy workbook (Year/Month/Week + Invoices-like col + Plant),
    or any flat sheet with those columns."""
    data = await file.read()
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    cyc = DB.select("cycle", {"id": v[0]["cycle_id"]})[0]
    q_months = {(cyc["quarter"] - 1) * 3 + i for i in (1, 2, 3)}
    aliases = DB.select("plant_alias")
    wb = load_wb(data)
    best = None
    for name in wb.sheetnames:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try: header = [str(h or "").strip().lower() for h in next(it)]
        except StopIteration: continue
        if "year" in header and "week" in header and any("plant" in h for h in header):
            best = (name, header); break
    if not best: raise HTTPException(400, "No flat sheet with Year/Week/Plant columns found in dummy file")
    name, header = best
    ws = wb[name]
    def find(*ns):
        for n in ns:
            for i, h in enumerate(header):
                if n in h: return i
        return None
    iY, iM, iW, iPl = find("year"), find("month"), find("week"), find("plant")
    iU = find("invoices", "units", "u")
    agg = defaultdict(float)
    it = ws.iter_rows(min_row=2, values_only=True)
    _empty = 0; _dc_cache = {}
    for r in it:
        if r is None or r[iY] in (None, ""):
            _empty += 1
            if _empty >= 1000: break
            continue
        _empty = 0
        try: mo, wk = int(r[iM]), int(r[iW])
        except Exception: continue
        if mo not in q_months: continue
        dc = plant_to_dc(r[iPl], aliases, _dc_cache)
        if dc: agg[(mo, wk, dc)] += num(r[iU]) if iU is not None else 1
    preview = {"sheet": name, "cells": len(agg),
               "by_dc": {}, "already_imported": False}
    for (mo, wk, dc), u in agg.items():
        preview["by_dc"][dc] = round(preview["by_dc"].get(dc, 0) + u)
    def do_commit():
        _clear_import("DUMMY", version_id=version_id)
        _pmap = ensure_periods_bulk(cyc["id"], {(a[0], a[1]) for a in agg.keys()})
        rows = [{"version_id": version_id, "cycle_id": cyc["id"], "period_id": _pmap[(mo, wk)]["id"],
                 "dc_code": dc, "channel_code": "", "kind": "DUMMY", "units": u}
                for (mo, wk, dc), u in agg.items()]
        DB.upsert("actual", rows, ["version_id", "period_id", "dc_code", "channel_code", "kind"])
        return {"rows": len(rows), "meta": preview}
    return _commit_or_preview(commit, "DUMMY", file.filename, preview, do_commit)

@app.post("/import/customers")
async def import_customers(file: UploadFile = File(...), version_id: int = Form(...), commit: bool = Form(False),
                           sheet: str = Form(""), authorization: Optional[str] = Header(None)):
    require_role(authorization, ('admin','planner'))
    """Customer-week forecast grid (Direct). Wide format: a header row containing customer names,
    rows keyed by week (col with 'M-W' labels or Week+Month columns). Unknown customers listed in preview;
    on commit they are auto-created (scalable) and flagged."""
    data = await file.read()
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    cyc = DB.select("cycle", {"id": v[0]["cycle_id"]})[0]
    q_months = {(cyc["quarter"] - 1) * 3 + i for i in (1, 2, 3)}
    wb = load_wb(data)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # find header row = row with >=3 non-numeric strings beyond col C
    hdr_i, hdr = None, None
    for i, r in enumerate(rows[:6]):
        strs = [c for c in (r or [])[3:] if isinstance(c, str) and c.strip()]
        if len(strs) >= 3: hdr_i, hdr = i, list(r); break
    if hdr is None: raise HTTPException(400, "Could not locate customer header row")
    SKIP = {"total", "direct", "total direct", "grand total", "sum",
            "italy", "china", "thailand", "total italy", "total china", "total thailand",
            "atlanta", "mexico", "meta llc",
            # Eleanor's subtotal / check columns (not customers)
            "ttl vas", "ttl no vas", "ttl ds", "ttl ds op", "vs lw", "month", "week",
            "meta/meu", "rw other", "ow other", "nuance", "rw", "ow", "+/- rw", "+/- ow",
            "addtl rw", "addtl ow", "ttl rw meta llc", "ttl ow meta llc",
            "atl/mx rw meta llc", "atl /mxow meta llc"}
    def _is_summary(name):
        n = str(name).strip().lower()
        if n in SKIP: return True
        # prefix/keyword guards for Eleanor's TTL/VS/META family
        if n.startswith("ttl ") or n.startswith("vs ") or n.startswith("+/-"): return True
        if n.startswith("meta/") or n.startswith("meta ") or "meta llc" in n: return True
        if n in ("rw other", "ow other", "rw", "ow", "nuance"): return True
        return False
    cust_cols = {j: str(c).strip() for j, c in enumerate(hdr)
                 if j >= 3 and isinstance(c, str) and c.strip() and not _is_summary(c)}
    known = {c["name"].strip().upper(): c for c in DB.select("customer")}
    alias_lookup = {}
    for c in DB.select("customer"):
        for a in (c.get("aliases") or []):
            alias_lookup[str(a).strip().upper()] = c
    grid = defaultdict(dict)
    for r in rows[hdr_i + 1:]:
        if not r: continue
        pk = None
        for cand in r[:3]:
            pk = parse_period_key(cand)
            if pk: break
        if not pk:
            # Week + Month numeric columns fallback (B=week, C=month like DC files)
            try:
                wk, mo = int(r[1]), int(r[2]); pk = (mo, wk)
            except Exception:
                continue
        mo, wk = pk
        if mo not in q_months: continue
        for j, name in cust_cols.items():
            if j < len(r): grid[(mo, wk)][name] = grid[(mo, wk)].get(name, 0) + num(r[j])
    names = sorted({n for g in grid.values() for n in g})
    unknown = [n for n in names if n.upper() not in known and n.upper() not in alias_lookup]
    preview = {"sheet": ws.title, "periods": len(grid), "customers": names, "new_customers": unknown,
               "total_units": round(sum(sum(g.values()) for g in grid.values())),
               "already_imported": _already_imported("CUSTFC", version_id=version_id)}
    def do_commit():
        _clear_import("CUSTFC", version_id=version_id)
        for n in unknown:
            add_customer({"name": n})
        cmap = {c["name"].strip().upper(): c["id"] for c in DB.select("customer")}
        for c in DB.select("customer"):
            for a in (c.get("aliases") or []): cmap[str(a).strip().upper()] = c["id"]
        out = []
        _pmap = ensure_periods_bulk(cyc["id"], grid.keys())
        for (mo, wk), g in grid.items():
            per = _pmap[(mo, wk)]
            for n, u in g.items():
                out.append({"version_id": version_id, "period_id": per["id"],
                            "customer_id": cmap[n.upper()], "units": u})
        DB.upsert("forecast_customer", out, ["version_id", "period_id", "customer_id"])
        return {"rows": len(out), "meta": preview}
    return _commit_or_preview(commit, "CUSTFC", file.filename, preview, do_commit)

# ------------------------------------------------------------------
# penetration / rates / overrides / dummy plan / meta llc
# ------------------------------------------------------------------
@app.get("/penetration")
def get_pen(version_id: int):
    return DB.select("penetration", {"version_id": version_id})

@app.post("/penetration")
def set_pen(payload: dict = Body(...)):
    rows = payload["rows"]
    for r in rows: r["version_id"] = payload["version_id"]
    DB.upsert("penetration", rows, ["version_id", "dc_code", "target_kind", "target_key"])
    return {"ok": True, "count": len(rows)}

def _recency_weighted(hist_rows, key_fn, val_fn, latest_ym):
    """Recency-weighted sums. weight = 0.85^(months_ago). Returns dict key -> weighted units."""
    out = defaultdict(float)
    ly, lm = latest_ym
    for h in hist_rows:
        age = (ly * 12 + lm) - (h["year"] * 12 + h["month_no"])
        if age < 0: continue
        w = 0.85 ** age
        out[key_fn(h)] += val_fn(h) * w
    return out

def _norm_cust_name(s):
    """Normalize customer names so BO actuals (DILLARDS, WALMART STORES INC) match Eleanor's
    customer table (DILLARD'S, WALMART STORES INC). Removes apostrophes + corporate suffixes."""
    import re as _re
    s = str(s or "").upper().strip()
    s = s.replace("\u2019", "").replace("'", "")          # remove apostrophes (no space)
    s = _re.sub(r"[,\.\-\(\)#/]", " ", s)                 # other punctuation -> space
    s = _re.sub(r"\b(INC|CORP|CORPORATION|LLC|LTD|CO|COMPANY|USA|US)\b", "", s)
    s = _re.sub(r"\s+", " ", s).strip()
    return s

@app.get("/customer_aliases")
def list_customer_aliases():
    aliases = DB.select("customer_alias")
    custs = {c["id"]: c["name"] for c in DB.select("customer")}
    return [{"id": a["id"], "actuals_name": a["actuals_name"],
             "customer_id": a["customer_id"], "customer_name": custs.get(a["customer_id"], "?")}
            for a in aliases]

@app.post("/customer_aliases")
def add_customer_alias(payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    require_role(authorization, ("admin", "planner"))
    an = (payload.get("actuals_name") or "").strip()
    cid = payload.get("customer_id")
    if not an or cid is None: raise HTTPException(400, "actuals_name and customer_id required")
    existing = [a for a in DB.select("customer_alias") if a["actuals_name"].strip().upper() == an.upper()]
    if existing:
        DB.update("customer_alias", {"id": existing[0]["id"]}, {"customer_id": cid})
        return {"ok": True, "updated": True}
    return DB.insert("customer_alias", {"actuals_name": an, "customer_id": cid})

@app.delete("/customer_aliases/{aid}")
def del_customer_alias(aid: int, authorization: Optional[str] = Header(None)):
    require_role(authorization, ("admin", "planner"))
    DB.delete("customer_alias", {"id": aid})
    return {"ok": True}

@app.get("/penetration/unmatched_actuals")
def unmatched_actuals(version_id: int):
    """BO actuals customer names (with export volume) that don't map to any customer or alias —
    so the user knows which aliases to add."""
    customers_all = DB.select("customer")
    by_norm = {}
    for c in customers_all:
        by_norm[_norm_cust_name(c["name"])] = c["id"]
        for a in (c.get("aliases") or []): by_norm[_norm_cust_name(a)] = c["id"]
    for al in DB.select("customer_alias"): by_norm[_norm_cust_name(al["actuals_name"])] = al["customer_id"]
    from collections import defaultdict as _dd
    exp = _dd(float)
    for a in DB.select("actual_customer", {"version_id": version_id}):
        if a["dc_code"] in ("IT", "CN", "TH"): exp[a["customer_name"]] += a["units"]
    out = []
    for nm, u in exp.items():
        n = _norm_cust_name(nm)
        if n in by_norm or n.startswith("COSTCO"): continue
        out.append({"actuals_name": nm, "export_units": round(u)})
    out.sort(key=lambda x: -x["export_units"])
    return {"unmatched": out[:50], "customers": [{"id": c["id"], "name": c["name"]} for c in customers_all]}

@app.get("/penetration/suggest")
def suggest_pen(version_id: int, months: int = 0):
    """Recency-weighted penetration from ALL imported history (0.85^months_ago).
    months=0 means use everything (recommended); a positive value caps the window.
    Channel: DC share of total per channel. Customer: country share of that customer.
    Never auto-applies — frontend previews drift, user clicks Apply."""
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    hist = [h for h in DB.select("history") if h["kind"] == "GROSS" and h["channel_code"]]
    basis = "recency-weighted over all imported history"
    if hist:
        latest = max((h["year"], h["month_no"]) for h in hist)
        if months > 0:
            lo = latest[0] * 12 + latest[1] - (months - 1)
            hist = [h for h in hist if h["year"] * 12 + h["month_no"] >= lo]
            basis = f"recency-weighted over last {months} months"
        tot = _recency_weighted(hist, lambda h: h["channel_code"], lambda h: h["units"], latest)
        per_dc = _recency_weighted(hist, lambda h: (h["dc_code"], h["channel_code"]), lambda h: h["units"], latest)
    else:
        basis = "current cycle QTD (no history imported yet)"
        acts = DB.select("actual", {"version_id": version_id})
        tot, per_dc = defaultdict(float), defaultdict(float)
        for a in acts:
            if a["kind"] != "GROSS" or not a["channel_code"]: continue
            tot[a["channel_code"]] += a["units"]; per_dc[(a["dc_code"], a["channel_code"])] += a["units"]
    out = [{"dc_code": dc, "target_kind": "CHANNEL", "target_key": ch, "pct": round(u / tot[ch], 4)}
           for (dc, ch), u in per_dc.items() if tot[ch] > 0]
    # --- customer -> country penetration, derived from actuals (with fallback ladder) ---
    # Fallback per customer, in order:
    #   1. customer's OWN export actuals (plant codes IT/CN/TH)
    #   2. brand-code mix: blend the customer's brand mix (from their actuals) x each brand's export country split
    #   3. collection-mix: same idea with OPTICAL/SUN/ELECTRONICS
    #   4. channel: same idea with the customer's channel
    #   5. overall Direct export split (last resort)
    # COSTCO variants share COSTCO WHOLESALE's split; user-managed customer_alias maps BO names -> customer ids.
    export_dcs = ("IT", "CN", "TH")
    customers_all = DB.select("customer")
    cust_by_name = {}
    for c in customers_all:
        cust_by_name[_norm_cust_name(c["name"])] = c["id"]
        for a in (c.get("aliases") or []): cust_by_name[_norm_cust_name(a)] = c["id"]
    # user alias table: actuals_name (normalized) -> customer_id
    for al in DB.select("customer_alias"):
        cust_by_name[_norm_cust_name(al["actuals_name"])] = al["customer_id"]
    # built-in COSTCO grouping: any Eleanor COSTCO variant <- actuals "COSTCO WHOLESALE"
    id_by_norm = {_norm_cust_name(c["name"]): c["id"] for c in customers_all}
    costco_ids = [cid for nm, cid in id_by_norm.items() if nm.startswith("COSTCO")]

    # helper to resolve a BO actuals customer name -> our customer id
    def _resolve(nm):
        n = _norm_cust_name(nm)
        if n in cust_by_name: return cust_by_name[n]
        if n.startswith("COSTCO"):   # COSTCO WHOLESALE(S) -> all COSTCO variants handled below
            return "COSTCO_GROUP"
        return None

    # 1) customer OWN export actuals (per-customer id) + COSTCO group bucket
    cust_rows = []
    costco_country = defaultdict(float)   # dc -> units (shared COSTCO split)
    costco_exp = 0.0
    for a in DB.select("actual_customer", {"version_id": version_id}):
        if a["dc_code"] not in export_dcs: continue
        cid = _resolve(a["customer_name"])
        if cid == "COSTCO_GROUP":
            costco_country[a["dc_code"]] += a["units"]; costco_exp += a["units"]
        elif cid is not None:
            cust_rows.append({"year": a["year"], "month_no": a["month_no"],
                              "dc_code": a["dc_code"], "customer_id": cid, "units": a["units"]})
    for h in DB.select("history_customer"):
        if h["dc_code"] in export_dcs:
            cust_rows.append({"year": h["year"], "month_no": h["month_no"],
                              "dc_code": h["dc_code"], "customer_id": h["customer_id"], "units": h["units"]})

    have_own = {}   # cid -> {dc: pct}
    if cust_rows:
        clatest = max((r["year"], r["month_no"]) for r in cust_rows)
        if months > 0:
            lo = clatest[0] * 12 + clatest[1] - (months - 1)
            cust_rows = [r for r in cust_rows if r["year"] * 12 + r["month_no"] >= lo]
        cexp = _recency_weighted(cust_rows, lambda r: r["customer_id"], lambda r: r["units"], clatest)
        ccountry = _recency_weighted(cust_rows, lambda r: (r["dc_code"], r["customer_id"]), lambda r: r["units"], clatest)
        for (dc, cid), u in ccountry.items():
            if cexp.get(cid, 0) > 0:
                have_own.setdefault(cid, {})[dc] = u / cexp[cid]

    # COSTCO shared split -> apply to all COSTCO variant ids
    if costco_exp > 0:
        for cid in costco_ids:
            have_own[cid] = {dc: costco_country.get(dc, 0.0) / costco_exp for dc in export_dcs}

    # 2-4) dimension -> country splits, and each customer's dimensional mix (from actuals)
    dimc = defaultdict(lambda: defaultdict(float))   # (kind,val) -> {dc: units}
    for d in DB.select("actual_dim_country", {"version_id": version_id}):
        dimc[(d["dim_kind"], d["dim_val"])][d["dc_code"]] += d["units"]
    dim_split = {}   # (kind,val) -> {dc: pct}
    for key, dd in dimc.items():
        tot = sum(dd.values())
        if tot > 0: dim_split[key] = {dc: dd.get(dc, 0.0) / tot for dc in export_dcs}
    # customer's dimensional mix (weights) from their own actuals (all DCs)
    cust_dim_mix = defaultdict(lambda: defaultdict(float))   # cid -> {(kind,val): units}
    for cd in DB.select("actual_cust_dim", {"version_id": version_id}):
        cid = _resolve(cd["customer_name"])
        if cid in (None, "COSTCO_GROUP"):
            if cid == "COSTCO_GROUP":
                for ccid in costco_ids:
                    cust_dim_mix[ccid][(cd["dim_kind"], cd["dim_val"])] += cd["units"]
            continue
        cust_dim_mix[cid][(cd["dim_kind"], cd["dim_val"])] += cd["units"]

    def _blend(cid, kind):
        """Blend a customer's mix of `kind` dims x each dim's export country split."""
        mix = {k: u for (k, v), u in [((k2, v2), u2) for (k2, v2), u2 in cust_dim_mix.get(cid, {}).items()] if k[0] == kind} if False else None
        # gather this customer's weights for the given dim kind
        weights = {(k, v): u for (k, v), u in cust_dim_mix.get(cid, {}).items() if k == kind}
        tot = sum(weights.values())
        if tot <= 0: return None
        acc = {dc: 0.0 for dc in export_dcs}
        used = 0.0
        for (k, v), u in weights.items():
            sp = dim_split.get((k, v))
            if not sp: continue
            for dc in export_dcs: acc[dc] += (u / tot) * sp[dc]
            used += u / tot
        if used <= 0: return None
        return {dc: acc[dc] / used for dc in export_dcs}   # renormalize over dims that had a split

    # 5) overall Direct export split (last resort)
    overall = defaultdict(float)
    for key, dd in dimc.items():
        if key[0] == "BRAND":
            for dc in export_dcs: overall[dc] += dd.get(dc, 0.0)
    ov_tot = sum(overall.values())
    overall_split = {dc: (overall[dc] / ov_tot if ov_tot > 0 else 1/3) for dc in export_dcs}

    cust_out = []
    tier_counts = defaultdict(int)
    for c in customers_all:
        cid = c["id"]
        split = have_own.get(cid); tier = "own"
        if not split: split = _blend(cid, "BRAND");   tier = "brand" if split else tier
        if not split: split = _blend(cid, "MIX");     tier = "mix" if split else tier
        if not split: split = _blend(cid, "CHANNEL"); tier = "channel" if split else tier
        if not split: split, tier = overall_split, "overall"
        split = {dc: float(split.get(dc, 0.0)) for dc in export_dcs}   # guarantee all keys
        s = sum(split.values())
        if s <= 0: continue
        tier_counts[tier] += 1
        for dc in export_dcs:
            cust_out.append({"dc_code": dc, "target_kind": "CUSTOMER", "target_key": str(cid),
                             "pct": round(split[dc] / s, 4)})

    cust_basis = ("customer->country from actuals: " +
                  ", ".join(f"{k}={v}" for k, v in sorted(tier_counts.items())) ) if tier_counts else "no customer actuals found"

    n_months = len({(h["year"], h["month_no"]) for h in hist}) if hist else 0
    return {"suggestions": out, "customer_suggestions": cust_out, "basis": basis,
            "customer_basis": cust_basis, "customer_tiers": dict(tier_counts),
            "history_loaded": bool(hist or cust_rows), "months_of_history": n_months}

@app.get("/rates")
def get_rates(version_id: int):
    return DB.select("rate", {"version_id": version_id})

@app.post("/rates")
def set_rates(payload: dict = Body(...)):
    rows = payload["rows"]
    for r in rows: r["version_id"] = payload["version_id"]
    DB.upsert("rate", rows, ["version_id", "dc_code", "kind"])
    return {"ok": True}

@app.get("/dummy_plan")
def get_dummy(version_id: int):
    return DB.select("dummy_plan", {"version_id": version_id})

@app.post("/dummy_plan")
def set_dummy(payload: dict = Body(...)):
    rows = payload["rows"]
    for r in rows: r["version_id"] = payload["version_id"]
    DB.upsert("dummy_plan", rows, ["version_id", "period_id", "dc_code"])
    return {"ok": True}

@app.get("/overrides")
def get_overrides(version_id: int):
    return [o for o in DB.select("override", {"version_id": version_id}) if o.get("active", True)]

@app.post("/overrides")
def add_override(payload: dict = Body(...)):
    if not (payload.get("reason") or "").strip():
        raise HTTPException(400, "reason required for overrides")
    payload.setdefault("active", True)
    payload["created_at"] = datetime.datetime.utcnow().isoformat()
    return DB.insert("override", payload)

@app.delete("/overrides/{oid}")
def del_override(oid: int):
    DB.update("override", {"id": oid}, {"active": False})
    return {"ok": True}

@app.post("/overrides/shift")
def shift_units(payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    """Move `units` from period_from -> period_to for a customer within a country (or a whole
    country total). Creates/updates overrides on both weeks so the move is auditable and the
    quarter total is preserved by construction. Additive: never deletes source data.
    payload: version_id, dc_code (IT/CN/TH), scope ('customer'|'country'),
             customer_id (if scope=customer), period_from, period_to, units, reason, author"""
    require_role(authorization, ("admin", "planner"))
    version_id = payload["version_id"]; dc = payload["dc_code"]
    scope = payload.get("scope", "customer"); units = float(payload["units"])
    pf = payload["period_from"]; pt = payload["period_to"]
    reason = (payload.get("reason") or "").strip()
    author = payload.get("author", "")
    if not reason: raise HTTPException(400, "reason required")
    if units <= 0: raise HTTPException(400, "units must be positive")
    if pf == pt: raise HTTPException(400, "from and to weeks must differ")
    if dc not in ("IT", "CN", "TH"): raise HTTPException(400, "shift only applies to export DCs (IT/CN/TH)")

    m = compute_model(version_id)
    act = set(m["actualized_periods"])
    if pf in act or pt in act:
        raise HTTPException(400, "cannot shift into or out of an actualized week")

    def cur_val(cid, pid):
        for r in m["customer_grid"]["countries"][dc]["rows"]:
            if r["period_id"] == pid:
                c = r["cells"].get(cid); return c["v"] if c else 0.0
        return 0.0

    targets = []
    if scope == "country":
        # move proportionally across every customer that has volume in the FROM week
        rowf = next((r for r in m["customer_grid"]["countries"][dc]["rows"] if r["period_id"] == pf), None)
        base = sum(max(c["v"], 0) for c in (rowf["cells"].values() if rowf else []))
        if base <= 0: raise HTTPException(400, "no positive volume in the from-week to move")
        for cid, c in (rowf["cells"].items() if rowf else []):
            if c["v"] > 0:
                targets.append((cid, units * (c["v"] / base)))
    else:
        cid = int(payload["customer_id"])
        avail = cur_val(cid, pf)
        if units > avail + 1e-6:
            raise HTTPException(400, f"cannot move {round(units)} — only {round(avail)} available in the from-week")
        targets = [(cid, units)]

    made = []
    for cid, mv in targets:
        from_new = cur_val(cid, pf) - mv
        to_new = cur_val(cid, pt) + mv
        for pid, val in ((pf, from_new), (pt, to_new)):
            existing = next((o for o in DB.select("override", {"version_id": version_id})
                             if o.get("active", True) and o["dc_code"] == dc
                             and o["target_kind"] == "CUSTOMER" and str(o["target_key"]) == str(cid)
                             and o["period_id"] == pid), None)
            body = {"version_id": version_id, "dc_code": dc, "target_kind": "CUSTOMER",
                    "target_key": str(cid), "period_id": pid, "units": round(val, 3),
                    "reason": f"Shift {round(mv)} {'in' if pid==pt else 'out'}: {reason}",
                    "author": author, "active": True,
                    "created_at": datetime.datetime.utcnow().isoformat()}
            if existing:
                DB.update("override", {"id": existing["id"]}, {"units": round(val, 3), "reason": body["reason"], "author": author})
            else:
                DB.insert("override", body)
        made.append({"customer_id": cid, "moved": round(mv)})
    return {"ok": True, "moved_total": round(units), "targets": made}

@app.get("/plan_baseline")
def plan_baseline(version_id: int):
    """Per-customer and per-country quarter totals from the ORIGINAL model (no overrides),
    so the UI can warn when hand edits drift the plan up or down."""
    m = compute_model(version_id, ignore_overrides=True)
    cust_tot = defaultdict(float); country_tot = defaultdict(float)
    for dc in ("IT", "CN", "TH"):
        for r in m["customer_grid"]["countries"][dc]["rows"]:
            for cid, c in r["cells"].items():
                cust_tot[f"{dc}:{cid}"] += c["v"]; country_tot[dc] += c["v"]
    return {"customer_totals": dict(cust_tot), "country_totals": dict(country_tot),
            "grand_total": round(sum(country_tot.values()))}

@app.post("/overrides/reallocate_negatives")
def reallocate_negatives(payload: dict = Body(...)):
    """One-click: for every negative export-DC customer-week (IT/CN/TH), create an override that
    zeroes it and logs the reason. Returns physically process through Atlanta, so a negative direct
    week doesn't belong on an export DC. Additive: creates override rows, never deletes data."""
    version_id = payload["version_id"]
    author = payload.get("author", "")
    m = compute_model(version_id)
    act_set = set(m["actualized_periods"])
    made = []
    existing = {(o["dc_code"], o["target_key"], o["period_id"]) for o in DB.select("override", {"version_id": version_id}) if o.get("active", True)}
    for dc in ("IT", "CN", "TH"):
        for row in m["customer_grid"]["countries"][dc]["rows"]:
            if row["period_id"] in act_set: continue   # never touch actuals
            for cid, cell in row["cells"].items():
                if cell["v"] < 0 and (dc, str(cid), row["period_id"]) not in existing:
                    cust = next((c["name"] for c in m["customers"] if c["id"] == cid), str(cid))
                    DB.insert("override", {"version_id": version_id, "dc_code": dc, "target_kind": "CUSTOMER",
                                           "target_key": str(cid), "period_id": row["period_id"], "units": 0,
                                           "reason": f"Negative reallocated to Atlanta (returns process through ATL) - was {round(cell['v'])}",
                                           "author": author, "active": True,
                                           "created_at": datetime.datetime.utcnow().isoformat()})
                    made.append({"dc": dc, "customer": cust, "week": row["label"], "was": round(cell["v"])})
    return {"ok": True, "count": len(made), "reallocated": made}

# ------------------------------------------------------------------
# model compute — the engine
# ------------------------------------------------------------------
def compute_model(version_id: int, ignore_overrides: bool = False):
    vrows = DB.select("version", {"id": version_id})
    if not vrows: raise HTTPException(404, "version not found")
    v = vrows[0]
    cyc = DB.select("cycle", {"id": v["cycle_id"]})[0]
    periods = sorted(DB.select("period", {"cycle_id": cyc["id"]}), key=lambda p: p["sort"])
    pid_order = [p["id"] for p in periods]
    channels = [c for c in DB.select("channel", order="sort") if c["active"]]
    ch_codes = [c["code"] for c in channels]
    dcs = [d for d in DB.select("dc", order="sort") if d["active"]]
    customers = [c for c in DB.select("customer", order="sort") if c["active"]]

    fc = {(r["period_id"], r["channel_code"]): r["units"] for r in DB.select("forecast_channel", {"version_id": version_id})}
    fcust = {(r["period_id"], r["customer_id"]): r["units"] for r in DB.select("forecast_customer", {"version_id": version_id})}
    acts = DB.select("actual", {"version_id": version_id})
    A = defaultdict(float)
    for a in acts: A[(a["period_id"], a["dc_code"], a["channel_code"], a["kind"])] += a["units"]
    per_sort = {p["id"]: p["sort"] for p in periods}
    with_gross = sorted({a["period_id"] for a in acts if a["kind"] == "GROSS" and a["period_id"] in per_sort},
                        key=lambda pid: per_sort[pid])
    # Eleanor's rule: actualize completed weeks only — the newest week with data is in progress.
    cutoff = get_setting(f"cutoff_v{version_id}", None)
    if cutoff is not None:
        actualized = [pid for pid in with_gross if per_sort[pid] <= int(cutoff)]
    else:
        actualized = with_gross[:-1] if len(with_gross) > 1 else with_gross
    act_set = set(actualized)

    pen = {(r["dc_code"], r["target_kind"], r["target_key"]): r["pct"]
           for r in DB.select("penetration", {"version_id": version_id})}
    rates = {(r["dc_code"], r["kind"]): r["pct"] for r in DB.select("rate", {"version_id": version_id})}
    dplan = {(r["period_id"], r["dc_code"]): r["units"] for r in DB.select("dummy_plan", {"version_id": version_id})}
    ovr = [] if ignore_overrides else [o for o in DB.select("override", {"version_id": version_id}) if o.get("active", True)]
    ovr_ch = {(o["dc_code"], o["target_key"], o["period_id"]): o for o in ovr if o["target_kind"] == "CHANNEL"}
    ovr_cu = {(o["dc_code"], o["target_key"], o["period_id"]): o for o in ovr if o["target_kind"] == "CUSTOMER"}

    def penv(dc, kind, key): return pen.get((dc, kind, key), 0.0)

    model = {"version": v, "cycle": cyc, "periods": periods, "channels": channels,
             "customers": customers, "actualized_periods": actualized, "dcs": dcs, "warnings": []}

    # --- channel-level DC grids (ATL, MX explicit; Direct = residual) ---
    grids = {}
    for dc in ["ATL", "MX"]:
        g = {"rows": [], "dc": dc}
        for p in periods:
            pid = p["id"]; row = {"period_id": pid, "label": p["label"], "actual": pid in act_set, "cells": {}}
            for ch in ch_codes:
                if pid in act_set:
                    val = A.get((pid, dc, ch, "GROSS"), 0.0)
                    src = "actual"
                else:
                    val = penv(dc, "CHANNEL", ch) * fc.get((pid, ch), 0.0)
                    src = "model"
                o = ovr_ch.get((dc, ch, pid))
                if o and pid not in act_set:
                    val = o["units"]; src = "override"
                row["cells"][ch] = {"v": val, "src": src}
            gross = sum(c["v"] for c in row["cells"].values())
            if pid in act_set:
                acc = A.get((pid, dc, "", "ACC"), 0.0)
                ret = sum(A.get((pid, dc, ch, "RETURN"), 0.0) for ch in ch_codes)
            else:
                acc = gross * rates.get((dc, "ACC"), 0.0)
                ret = gross * rates.get((dc, "RET"), 0.0)
            dummy = A.get((pid, dc, "", "DUMMY"), 0.0) if pid in act_set else dplan.get((pid, dc), 0.0)
            rw = A.get((pid, dc, "", "RW"), 0.0) if pid in act_set else penv(dc, "LINE", "RW") * (fc.get((pid, "RW_TTL"), 0.0) - fc.get((pid, "RW_META"), 0.0))
            nuance = A.get((pid, dc, "", "NUANCE"), 0.0) if pid in act_set else penv(dc, "LINE", "NUANCE") * fc.get((pid, "NUANCE"), 0.0)
            ow = A.get((pid, dc, "", "OW"), 0.0) if pid in act_set else penv(dc, "LINE", "OW") * (fc.get((pid, "OW_TTL"), 0.0) - fc.get((pid, "OW_META"), 0.0))
            row.update({"gross": gross, "acc": acc, "ret": ret, "ttl": gross + acc + ret,
                        "dummy": dummy, "rw": rw, "nuance": nuance, "ow": ow})
            g["rows"].append(row)
        grids[dc] = g

    # Direct residual by channel
    g = {"rows": [], "dc": "DIRECT"}
    for i, p in enumerate(periods):
        pid = p["id"]; row = {"period_id": pid, "label": p["label"], "actual": pid in act_set, "cells": {}}
        for ch in ch_codes:
            if pid in act_set:
                val = sum(A.get((pid, d, ch, "GROSS"), 0.0) for d in ("IT", "CN", "TH"))
                src = "actual"
            else:
                val = fc.get((pid, ch), 0.0) - grids["ATL"]["rows"][i]["cells"][ch]["v"] - grids["MX"]["rows"][i]["cells"][ch]["v"]
                src = "model"
                if val < 0:
                    model["warnings"].append(f"Direct negative: {ch} {p['label']} = {round(val)}")
            row["cells"][ch] = {"v": val, "src": src}
        row["gross"] = sum(c["v"] for c in row["cells"].values())
        meta_rw = A.get((pid, "CN", "", "META_RW"), 0.0) + A.get((pid, "IT", "", "META_RW"), 0.0) + A.get((pid, "TH", "", "META_RW"), 0.0) \
            if pid in act_set else fc.get((pid, "RW_META"), 0.0)
        meta_ow = A.get((pid, "CN", "", "META_OW"), 0.0) + A.get((pid, "IT", "", "META_OW"), 0.0) + A.get((pid, "TH", "", "META_OW"), 0.0) \
            if pid in act_set else fc.get((pid, "OW_META"), 0.0)
        row["meta_rw"], row["meta_ow"], row["meta_ttl"] = meta_rw, meta_ow, meta_rw + meta_ow
        g["rows"].append(row)
    grids["DIRECT"] = g

    # --- customer base: imported Direct file if present, otherwise DERIVED from actuals ---
    # The app's purpose is to replace Eleanor's Direct workbook, so the file must be optional.
    # Derived mode uses this week's own BO actuals: each customer's recency-weighted share of
    # export (IT/CN/TH) volume, applied to the Direct residual, and split by country using that
    # customer's own actual country mix (saved penetration still wins when present).
    _direct_by_pid = {r["period_id"]: r["gross"] for r in grids["DIRECT"]["rows"]}
    derived_base = not fcust
    cust_share, cust_split, act_cust = {}, {}, {}
    if derived_base:
        _byname = {}
        for c in customers:
            _byname[_norm_cust_name(c["name"])] = c["id"]
            for a in (c.get("aliases") or []): _byname[_norm_cust_name(a)] = c["id"]
        for al in DB.select("customer_alias"): _byname[_norm_cust_name(al["actuals_name"])] = al["customer_id"]
        _costco = [cid for nm, cid in _byname.items() if nm.startswith("COSTCO")]
        pid_by_mw = {(p["month_no"], p["week_no"]): p["id"] for p in periods}
        rows_ac = DB.select("actual_customer", {"version_id": version_id})
        latest = max(((r["year"], r["month_no"]) for r in rows_ac), default=None)
        exp_by_cid, exp_by_cid_dc, exp_total = defaultdict(float), defaultdict(float), 0.0
        for r in rows_ac:
            if r["dc_code"] not in ("IT", "CN", "TH"): continue
            u = r["units"] or 0.0
            w = (0.85 ** max(0, (latest[0] * 12 + latest[1]) - (r["year"] * 12 + r["month_no"]))) if latest else 1.0
            exp_total += u * w
            n = _norm_cust_name(r["customer_name"])
            cids = _costco if (n.startswith("COSTCO") and n not in _byname) else \
                   ([_byname[n]] if n in _byname else [])
            if not cids: continue
            portion = (u * w) / len(cids)
            for cid in cids:
                exp_by_cid[cid] += portion
                exp_by_cid_dc[(cid, r["dc_code"])] += portion
            pid = pid_by_mw.get((r["month_no"], r["week_no"]))
            if pid is not None:
                for cid in cids:
                    act_cust[(pid, cid, r["dc_code"])] = act_cust.get((pid, cid, r["dc_code"]), 0.0) + (u / len(cids))
        if exp_total > 0:
            for cid, v in exp_by_cid.items(): cust_share[cid] = v / exp_total
        # each customer's own country mix from actuals; overall export mix as the fallback
        overall = {dc: 0.0 for dc in ("IT", "CN", "TH")}
        for (cid, dc), v in exp_by_cid_dc.items(): overall[dc] += v
        ov_t = sum(overall.values())
        overall_split = {dc: (overall[dc] / ov_t if ov_t > 0 else 1 / 3) for dc in ("IT", "CN", "TH")}
        for cid, tot in exp_by_cid.items():
            if tot > 0:
                cust_split[cid] = {dc: exp_by_cid_dc.get((cid, dc), 0.0) / tot for dc in ("IT", "CN", "TH")}
        overall_split_ref = overall_split
        model["customer_source"] = {
            "mode": "derived_from_actuals",
            "customers_with_export_history": len(cust_share),
            "note": "No Direct file imported — customer split derived from this week's BO actuals."}
    else:
        overall_split_ref = {"IT": 1/3, "CN": 1/3, "TH": 1/3}
        model["customer_source"] = {"mode": "imported_direct_file",
                                    "note": "Customer weekly totals came from the imported Direct workbook."}

    # --- customer-level Direct + country split ---
    cust_grid = {"rows": [], "countries": {}}
    for dc in ["IT", "CN", "TH"]:
        cust_grid["countries"][dc] = {"rows": []}
    for p in periods:
        pid = p["id"]
        row = {"period_id": pid, "label": p["label"], "actual": pid in act_set, "cells": {}}
        crow = {dc: {"period_id": pid, "label": p["label"], "actual": pid in act_set, "cells": {}, "total": 0.0}
                for dc in ("IT", "CN", "TH")}
        for cu in customers:
            base = fcust.get((pid, cu["id"]), 0.0)
            fixed_by_dc = None
            if derived_base:
                if pid in act_set:
                    # completed week: use what actually shipped, by country, straight from actuals
                    fixed_by_dc = {dc: act_cust.get((pid, cu["id"], dc), 0.0) for dc in ("IT", "CN", "TH")}
                    base = sum(fixed_by_dc.values())
                else:
                    base = cust_share.get(cu["id"], 0.0) * _direct_by_pid.get(pid, 0.0)
            country_sum = 0.0     # ALL tab = sum of the country-allocated values (post-override)
            any_override = False
            for dc in ("IT", "CN", "TH"):
                if fixed_by_dc is not None:
                    val = fixed_by_dc[dc]
                else:
                    pct = penv(dc, "CUSTOMER", str(cu["id"]))
                    if derived_base and pct == 0.0:
                        # no saved customer penetration yet: fall back to that customer's own
                        # actual country mix so the view works with zero manual setup
                        pct = (cust_split.get(cu["id"]) or overall_split_ref)[dc]
                    val = pct * base
                o = ovr_cu.get((dc, str(cu["id"]), pid))
                src = "model"
                if o: val, src = o["units"], "override"; any_override = True
                if val < 0:
                    model["warnings"].append(f"Negative {dc} {cu['name']} {p['label']} = {round(val)} — reclass to Atlanta?")
                crow[dc]["cells"][cu["id"]] = {"v": val, "src": src}
                crow[dc]["total"] += val
                country_sum += val
            # ALL tab reflects the reallocated country totals, not the raw base
            row["cells"][cu["id"]] = country_sum
        row["total"] = sum(row["cells"].values())
        cust_grid["rows"].append(row)
        for dc in ("IT", "CN", "TH"):
            if pid in act_set:
                crow[dc]["actual_total"] = sum(A.get((pid, dc, ch, "GROSS"), 0.0) for ch in ch_codes)
            cust_grid["countries"][dc]["rows"].append(crow[dc])

    # --- Direct reconciliation: Other (unnamed) bucket ties customer bottom-up to channel residual ---
    direct_gross_by_pid = {r["period_id"]: r["gross"] for r in grids["DIRECT"]["rows"]}
    for i, row in enumerate(cust_grid["rows"]):
        pid = row["period_id"]
        # "named" = what actually lands in the country tabs (customer % x base), summed over IT/CN/TH.
        country_named = {dc: cust_grid["countries"][dc]["rows"][i]["total"] for dc in ("IT", "CN", "TH")}
        named = sum(country_named.values())
        resid = direct_gross_by_pid.get(pid, 0.0)
        other = resid - named
        row["base_customer_total"] = row["total"]   # raw planner bottom-up (kept for reference)
        row["direct_residual"] = resid
        row["named_total"] = named
        row["other"] = max(other, 0.0)               # export DCs never negative
        row["other_raw"] = other                      # keep the true plug for reconciliation/diagnostics
        row["reconciled_total"] = named + row["other"]
        base = named
        # Export DCs never carry negatives (Eleanor's rule: returns only through Atlanta).
        # Floor the Other plug at 0 on country views; a negative plug means named customers
        # over-allocated that week — that shortfall belongs to Atlanta, not an export DC.
        other_export = max(other, 0.0)
        for dc in ("IT", "CN", "TH"):
            share = (country_named[dc] / base) if base > 0 else (1 / 3)
            crow = cust_grid["countries"][dc]["rows"][i]
            crow["other"] = other_export * share
            crow["reconciled_total"] = crow["total"] + crow["other"]
        if resid > 0 and abs(other) > 0.15 * resid:
            floored = " (floored to 0 on export DCs — over-allocation stays in Atlanta)" if other < 0 else ""
            model["warnings"].append(
                f"Direct mix gap: {row['label']} allocated {round(named)} vs residual {round(resid)} "
                f"(Other {round(other):+d}){floored} — review customer list/timing/%")

    model["grids"] = grids
    model["customer_grid"] = cust_grid

    # --- capacity lens (DC view): workload = gross + accessories + returns + dummies; per working day ---
    wd = {p["id"]: (p.get("working_days") or 5) for p in periods}
    cap = {}
    for dc in ("ATL", "MX"):
        rows_c, work_vals = [], []
        for r in grids[dc]["rows"]:
            workload = r["gross"] + r["acc"] + r["ret"] + r["dummy"]
            rows_c.append({"period_id": r["period_id"], "label": r["label"], "actual": r["actual"],
                           "workload": workload, "per_day": workload / max(wd[r["period_id"]], 1),
                           "working_days": wd[r["period_id"]]})
            if not r["actual"]: work_vals.append(workload)
        med = sorted(work_vals)[len(work_vals) // 2] if work_vals else 0
        for rc in rows_c:
            rc["peak"] = (not rc["actual"]) and med > 0 and rc["workload"] > 1.2 * med
        cap[dc] = {"rows": rows_c, "median_fc_workload": med}
    for dc in ("IT", "CN", "TH"):
        rows_c = []
        for r in cust_grid["countries"][dc]["rows"]:
            g = (r.get("actual_total") if r["actual"] and r.get("actual_total") is not None else r["total"])
            rows_c.append({"period_id": r["period_id"], "label": r["label"], "actual": r["actual"],
                           "workload": g, "per_day": g / max(wd[r["period_id"]], 1),
                           "working_days": wd[r["period_id"]], "peak": False})
        cap[dc] = {"rows": rows_c}
    model["capacity"] = cap

    # --- summaries: monthly, QTD/to-go, totals ---
    def summarize(rows, keys):
        months, total = defaultdict(lambda: defaultdict(float)), defaultdict(float)
        qtd, togo = defaultdict(float), defaultdict(float)
        per_by_id = {p["id"]: p for p in periods}
        for r in rows:
            mo = per_by_id[r["period_id"]]["month_no"]
            for k in keys:
                val = r.get(k, 0.0)
                months[mo][k] += val; total[k] += val
                (qtd if r["actual"] else togo)[k] += val
        return {"months": {m: dict(vv) for m, vv in months.items()}, "total": dict(total),
                "qtd": dict(qtd), "togo": dict(togo)}
    keys = ["gross", "acc", "ret", "ttl", "dummy", "rw", "nuance", "ow"]
    model["summary"] = {dc: summarize(grids[dc]["rows"], keys if dc != "DIRECT" else ["gross", "meta_rw", "meta_ow", "meta_ttl"])
                        for dc in grids}
    # Direct reconciliation summary
    model["direct_recon"] = {
        "named_total": round(sum(r["named_total"] for r in cust_grid["rows"])),
        "residual_total": round(sum(r["direct_residual"] for r in cust_grid["rows"])),
        "other_total": round(sum(r["other"] for r in cust_grid["rows"])),          # floored (export view)
        "other_raw_total": round(sum(r.get("other_raw", r["other"]) for r in cust_grid["rows"])),
        "floored_to_atlanta": round(sum(r["other"] - r.get("other_raw", r["other"]) for r in cust_grid["rows"])),
    }
    for dc in ("IT", "CN", "TH"):
        model["summary"][dc] = summarize(
            [{"period_id": r["period_id"], "actual": r["actual"],
              "gross": (r.get("actual_total") if r["actual"] and r.get("actual_total") is not None
                        else r.get("reconciled_total", r["total"]))}
             for r in cust_grid["countries"][dc]["rows"]], ["gross"])
    return model

def slim_model(m):
    """JSON-safe rounded copy for the frontend."""
    def rr(x): return round(x, 1) if isinstance(x, float) else x
    out = json.loads(json.dumps(m, default=str))
    return out

@app.get("/model")
def model_endpoint(version_id: int):
    return JSONResponse(slim_model(compute_model(version_id)))

@app.get("/model/vs")
def model_vs(version_id: int, prev_version_id: int):
    cur, prev = compute_model(version_id), compute_model(prev_version_id)
    out = {}
    for dc in cur["summary"]:
        c = cur["summary"][dc]["total"]; p = prev["summary"].get(dc, {}).get("total", {})
        out[dc] = {k: round(c.get(k, 0) - p.get(k, 0)) for k in c}
    per = {}
    for dc in ("ATL", "MX", "DIRECT"):
        prev_rows = {r["label"]: r for r in prev["grids"][dc]["rows"]}
        per[dc] = [{"label": r["label"], "delta": round(r["gross"] - prev_rows.get(r["label"], {}).get("gross", 0))}
                   for r in cur["grids"][dc]["rows"]]
    # decomposition on forecast periods: NA change vs penetration change vs overrides/actualization residual
    fc_c = {(r["period_id"], r["channel_code"]): r["units"] for r in DB.select("forecast_channel", {"version_id": version_id})}
    fc_p = {(r["period_id"], r["channel_code"]): r["units"] for r in DB.select("forecast_channel", {"version_id": prev_version_id})}
    pen_c = {(r["dc_code"], r["target_key"]): r["pct"] for r in DB.select("penetration", {"version_id": version_id}) if r["target_kind"] == "CHANNEL"}
    pen_p = {(r["dc_code"], r["target_key"]): r["pct"] for r in DB.select("penetration", {"version_id": prev_version_id}) if r["target_kind"] == "CHANNEL"}
    label_by_pid = {p2["id"]: p2["label"] for p2 in cur["periods"]}
    prev_pid_by_label = {p2["label"]: p2["id"] for p2 in prev["periods"]}
    decomp = {}
    ch_codes = [c2["code"] for c2 in cur["channels"]]
    for dc in ("ATL", "MX"):
        na_eff = pct_eff = 0.0
        for r in cur["grids"][dc]["rows"]:
            if r["actual"]: continue
            pid = r["period_id"]; ppid = prev_pid_by_label.get(label_by_pid.get(pid, ""), pid)
            for ch in ch_codes:
                fn, fo = fc_c.get((pid, ch), 0.0), fc_p.get((ppid, ch), 0.0)
                pn, po = pen_c.get((dc, ch), 0.0), pen_p.get((dc, ch), 0.0)
                na_eff += po * (fn - fo)
                pct_eff += (pn - po) * fn
        total_delta = out.get(dc, {}).get("gross", 0)
        decomp[dc] = {"na_change": round(na_eff), "pct_change": round(pct_eff),
                      "overrides_and_actualization": round(total_delta - na_eff - pct_eff),
                      "total": total_delta}
    return {"totals": out, "per_period": per, "decomposition": decomp}

@app.get("/monday_note")
def monday_note(version_id: int, prev_version_id: Optional[int] = None):
    """Email-ready Monday summary for the DCs."""
    m = compute_model(version_id)
    S = m["summary"]
    lines = [f"DC shipping forecast — {m['cycle']['label']} · {m['version']['week_tag']}",
             f"Actualized through {m['periods'][len(m['actualized_periods'])-1]['label'] if m['actualized_periods'] else '—'} (completed weeks)."]
    for dc, nm in [("ATL", "Atlanta"), ("MX", "Mexico"), ("DIRECT", "Direct (IT/CN/TH)")]:
        t = S[dc]["total"].get("gross", 0); q = S[dc]["qtd"].get("gross", 0); g = S[dc]["togo"].get("gross", 0)
        lines.append(f"{nm}: quarter {round(t):,} gross · QTD {round(q):,} · to go {round(g):,}")
    if prev_version_id:
        vs = model_vs(version_id, prev_version_id)
        for dc, nm in [("ATL", "Atlanta"), ("MX", "Mexico"), ("DIRECT", "Direct")]:
            d = vs["totals"].get(dc, {}).get("gross", 0)
            arrow = "up" if d >= 0 else "down"
            lines.append(f"vs last week — {nm}: {arrow} {abs(d):,}")
    ovr = [o for o in DB.select("override", {"version_id": version_id}) if o.get("active", True)]
    if ovr:
        lines.append("Notable adjustments: " + "; ".join(f"{o['dc_code']} {o['target_key']} — {o['reason']}" for o in ovr[:5]))
    if m["warnings"]:
        lines.append(f"Open checks: {len(m['warnings'])} (see app dashboard).")
    return {"text": "\n".join(lines)}

@app.get("/accuracy")
def accuracy(version_id: int):
    """WAPE of this version's model vs actuals on actualized periods (per DC, weekly + monthly)."""
    m = compute_model(version_id)
    fc = {(r["period_id"], r["channel_code"]): r["units"]
          for r in DB.select("forecast_channel", {"version_id": version_id})}
    pen = {(r["dc_code"], r["target_key"]): r["pct"]
           for r in DB.select("penetration", {"version_id": version_id}) if r["target_kind"] == "CHANNEL"}
    out = {}
    for dc in ("ATL", "MX"):
        abs_err = tot = 0.0
        for r in m["grids"][dc]["rows"]:
            if not r["actual"]: continue
            pid = r["period_id"]
            pred = sum(pen.get((dc, ch["code"]), 0) * fc.get((pid, ch["code"]), 0) for ch in m["channels"])
            act = r["gross"]
            abs_err += abs(pred - act); tot += act
        out[dc] = {"wape": round(abs_err / tot, 4) if tot else None, "weeks": len(m["actualized_periods"])}
    return out

# ------------------------------------------------------------------
# export — DC-familiar workbooks
# ------------------------------------------------------------------
@app.get("/export/dc")
def export_dc(version_id: int, which: str = Query("atl_mx", pattern="^(atl_mx|direct)$")):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    m = compute_model(version_id)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    GRAY = PatternFill("solid", fgColor="D9D9D9")
    ORANGE = PatternFill("solid", fgColor="FCE4CC")
    HDR = Font(bold=True); NAVY = Font(bold=True, color="1F3864")
    NUMFMT = "#,##0"        # commas, zero decimals
    PCTFMT = "0.0%"
    pct_lbl = f"{m['cycle']['label']} Forecast - {m['version']['week_tag']}"
    MN = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    def numcell(ws, r, c, val, fill=None, fmt=NUMFMT, bold=False):
        cell = ws.cell(r, c, round(val) if val is not None else None)
        cell.number_format = fmt
        if fill: cell.fill = fill
        if bold: cell.font = HDR
        return cell

    def channel_sheet(ws, dc):
        chs = m["channels"]
        ws["A1"] = pct_lbl; ws["A1"].font = NAVY
        tail = ["", "Accessories", "Returns", "Ttl Gross", "Acc % Gross", "Ret % Gross",
                "Dummies", "RW OTHR", "NUANCE", "OW OTHR"] if dc != "DIRECT" else \
               ["", "", "", "Ttl Gross", "", "", "", "META RW", "", "META OW"]
        hdrs = ["Week", "Month", f"Total {dict(ATL='Atlanta', MX='Mexico', DIRECT='Direct')[dc]}"] + \
               [c["name"] for c in chs] + tail
        for j, h in enumerate(hdrs, 1):
            cell = ws.cell(row=2, column=j, value=h); cell.font = HDR
        r0 = 3
        for i, row in enumerate(m["grids"][dc]["rows"]):
            mo, wk = row["label"].split("-")
            rr = r0 + i
            ws.cell(rr, 1, int(wk)); ws.cell(rr, 2, int(mo))
            numcell(ws, rr, 3, row["gross"])
            for j, c in enumerate(chs, 4):
                fill = GRAY if row["actual"] else (ORANGE if row["cells"][c["code"]]["src"] == "override" else None)
                numcell(ws, rr, j, row["cells"][c["code"]]["v"], fill)
            base = 4 + len(chs) + 1
            if dc == "DIRECT":
                vals = [(None, NUMFMT), (None, NUMFMT), (row["gross"], NUMFMT), (None, NUMFMT), (None, NUMFMT),
                        (None, NUMFMT), (row.get("meta_rw", 0), NUMFMT), (None, NUMFMT), (row.get("meta_ow", 0), NUMFMT)]
            else:
                q = row["gross"] + row["acc"] + row["ret"]
                vals = [(row["acc"], NUMFMT), (row["ret"], NUMFMT), (q, NUMFMT),
                        (row["acc"] / q if q else 0, PCTFMT), (row["ret"] / q if q else 0, PCTFMT),
                        (row["dummy"], NUMFMT), (row["rw"], NUMFMT), (row["nuance"], NUMFMT), (row["ow"], NUMFMT)]
            for j, (vv, fmt) in enumerate(vals):
                if vv is not None:
                    if fmt == PCTFMT:
                        cell = ws.cell(rr, base + j, round(vv, 4)); cell.number_format = fmt
                        if row["actual"]: cell.fill = GRAY
                    else:
                        numcell(ws, rr, base + j, vv, GRAY if row["actual"] else None)
        tr = r0 + len(m["grids"][dc]["rows"])
        c = ws.cell(tr, 3, f"=SUM(C{r0}:C{tr-1})"); c.font = HDR; c.number_format = NUMFMT
        for j in range(4, 4 + len(chs)):
            L2 = get_column_letter(j)
            c = ws.cell(tr, j, f"=SUM({L2}{r0}:{L2}{tr-1})"); c.font = HDR; c.number_format = NUMFMT
        sr = tr + 2
        ws.cell(sr - 1, 1, "Monthly").font = NAVY
        months = sorted(m["summary"][dc]["months"])
        for k, mo in enumerate(months):
            ws.cell(sr + k, 1, MN[mo])
            numcell(ws, sr + k, 3, m["summary"][dc]["months"][mo].get("gross", 0))
        ws.cell(sr + len(months), 1, "QTD actual").font = HDR
        numcell(ws, sr + len(months), 3, m["summary"][dc]["qtd"].get("gross", 0), bold=True)
        ws.cell(sr + len(months) + 1, 1, "To go").font = HDR
        numcell(ws, sr + len(months) + 1, 3, m["summary"][dc]["togo"].get("gross", 0), bold=True)
        ws.freeze_panes = "C3"
        for col in range(1, 4 + len(chs) + 12): ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions["A"].width = 7; ws.column_dimensions["B"].width = 7

    def customer_sheet(ws, dc):
        cg = m["customer_grid"]; custs = m["customers"]
        ws["A1"] = pct_lbl; ws["A1"].font = NAVY
        from openpyxl.utils import get_column_letter as gl
        oc = 4 + len(custs)
        hdrs = ["Week", "Month", f"{dict(IT='Italy', CN='China', TH='Thailand', ALL='Total')[dc]}"] + \
               [c["name"] for c in custs] + ["Other"]
        for j, h in enumerate(hdrs, 1): ws.cell(2, j, h).font = HDR
        rows = cg["rows"] if dc == "ALL" else cg["countries"][dc]["rows"]
        for i, row in enumerate(rows):
            mo, wk = row["label"].split("-")
            rr = 3 + i
            ws.cell(rr, 1, int(wk)); ws.cell(rr, 2, int(mo))
            if dc == "ALL":
                numcell(ws, rr, 3, row.get("reconciled_total", row["total"]))
                for j, cu in enumerate(custs, 4):
                    numcell(ws, rr, j, row["cells"].get(cu["id"], 0), GRAY if row["actual"] else None)
            else:
                _tot = (row.get("actual_total") if row["actual"] and row.get("actual_total") is not None
                        else row.get("reconciled_total", row["total"]))
                numcell(ws, rr, 3, _tot)
                for j, cu in enumerate(custs, 4):
                    c = row["cells"].get(cu["id"], {"v": 0, "src": "model"})
                    fill = GRAY if row["actual"] else (ORANGE if c["src"] == "override" else None)
                    numcell(ws, rr, j, c["v"], fill)
            numcell(ws, rr, oc, row.get("other", 0), GRAY if row["actual"] else None)
        tr = 3 + len(rows)
        for j in range(3, 5 + len(custs)):
            c = ws.cell(tr, j, f"=SUM({gl(j)}3:{gl(j)}{tr-1})"); c.font = HDR; c.number_format = NUMFMT
        # note WITHOUT leading '=' (that caused #NAME?); put as plain text with a colon
        ws.cell(1, oc, "Other: ties to channel Direct residual").font = Font(italic=True, color="595959")
        ws.freeze_panes = "D3"
        for col in range(1, 6 + len(custs)): ws.column_dimensions[gl(col)].width = 13
        ws.column_dimensions["A"].width = 7; ws.column_dimensions["B"].width = 7

    def total_sheet(ws, mode):
        """A combined Total sheet. atl_mx: monthly + Q totals per DC + returns/acc/dummy.
        direct: monthly + Q totals per country + Other."""
        ws["A1"] = pct_lbl + "  -  TOTAL"; ws["A1"].font = NAVY
        S = m["summary"]
        if mode == "atl_mx":
            dcs = [("ATL", "Atlanta"), ("MX", "Mexico"), ("DIRECT", "Direct")]
            cols = ["Month", "Atlanta", "Mexico", "Direct", "Grand total"]
            for j, h in enumerate(cols, 1): ws.cell(2, j, h).font = HDR
            months = sorted({mo for dc, _ in dcs for mo in S[dc]["months"]})
            r = 3
            for mo in months:
                ws.cell(r, 1, MN[mo])
                vals = [S[dc]["months"].get(mo, {}).get("gross", 0) for dc, _ in dcs]
                for j, v in enumerate(vals, 2): numcell(ws, r, j, v)
                numcell(ws, r, 5, sum(vals), bold=True); r += 1
            ws.cell(r, 1, "Q total").font = HDR
            qvals = [S[dc]["total"].get("gross", 0) for dc, _ in dcs]
            for j, v in enumerate(qvals, 2): numcell(ws, r, j, v, bold=True)
            numcell(ws, r, 5, sum(qvals), bold=True); r += 2
            # QTD / to-go
            ws.cell(r, 1, "QTD actual").font = HDR
            for j, (dc, _) in enumerate(dcs, 2): numcell(ws, r, j, S[dc]["qtd"].get("gross", 0))
            r += 1
            ws.cell(r, 1, "To go").font = HDR
            for j, (dc, _) in enumerate(dcs, 2): numcell(ws, r, j, S[dc]["togo"].get("gross", 0))
            r += 2
            # product lines total (ATL+MX)
            ws.cell(r, 1, "Product lines (ATL+MX)").font = NAVY; r += 1
            for lbl, key in [("Accessories", "acc"), ("Returns", "ret"), ("Dummies", "dummy"),
                             ("RW", "rw"), ("Nuance", "nuance"), ("OW", "ow")]:
                ws.cell(r, 1, lbl)
                numcell(ws, r, 2, S["ATL"]["total"].get(key, 0))
                numcell(ws, r, 3, S["MX"]["total"].get(key, 0))
                numcell(ws, r, 5, S["ATL"]["total"].get(key, 0) + S["MX"]["total"].get(key, 0), bold=True)
                r += 1
        else:
            countries = [("IT", "Italy"), ("CN", "China"), ("TH", "Thailand")]
            cols = ["Month", "Italy", "China", "Thailand", "Total Direct"]
            for j, h in enumerate(cols, 1): ws.cell(2, j, h).font = HDR
            months = sorted({mo for dc, _ in countries for mo in S[dc]["months"]})
            r = 3
            for mo in months:
                ws.cell(r, 1, MN[mo])
                vals = [S[dc]["months"].get(mo, {}).get("gross", 0) for dc, _ in countries]
                for j, v in enumerate(vals, 2): numcell(ws, r, j, v)
                numcell(ws, r, 5, sum(vals), bold=True); r += 1
            ws.cell(r, 1, "Q total").font = HDR
            qvals = [S[dc]["total"].get("gross", 0) for dc, _ in countries]
            for j, v in enumerate(qvals, 2): numcell(ws, r, j, v, bold=True)
            numcell(ws, r, 5, sum(qvals), bold=True); r += 2
            rec = m.get("direct_recon", {})
            ws.cell(r, 1, "Reconciliation").font = NAVY; r += 1
            for lbl, key in [("Named customers", "named_total"), ("Other (unnamed)", "other_total"),
                             ("Channel Direct residual", "residual_total")]:
                ws.cell(r, 1, lbl); numcell(ws, r, 5, rec.get(key, 0), bold=True); r += 1
        for col in range(1, 6): ws.column_dimensions[get_column_letter(col)].width = 16

    if which == "atl_mx":
        channel_sheet(wb.create_sheet("Atlanta"), "ATL")
        channel_sheet(wb.create_sheet("Mexico"), "MX")
        channel_sheet(wb.create_sheet("Direct"), "DIRECT")
        total_sheet(wb.create_sheet("Total"), "atl_mx")
        fname = f"Atlanta Mexico {m['cycle']['label']} {m['version']['week_tag']}.xlsx"
    else:
        customer_sheet(wb.create_sheet("Italy"), "IT")
        customer_sheet(wb.create_sheet("China"), "CN")
        customer_sheet(wb.create_sheet("Thailand"), "TH")
        customer_sheet(wb.create_sheet("Total Direct"), "ALL")
        total_sheet(wb.create_sheet("Total"), "direct")
        fname = f"Direct {m['cycle']['label']} {m['version']['week_tag']}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@app.post("/version/{vid}/cutoff")
def set_cutoff(vid: int, payload: dict = Body(...)):
    """Set/clear actualization cutoff (period sort value). null = auto (drop in-progress week)."""
    set_setting(f"cutoff_v{vid}", payload.get("cutoff"))
    return {"ok": True}


@app.get("/auth/status")
def auth_status():
    """Diagnostic: does the auth layer work? Confirms tables reachable + admin exists."""
    try:
        users = DB.select("app_user")
        return {"ok": True, "user_count": len(users),
                "admin_exists": any(u["username"] == "admin" for u in users),
                "db_mode": DB_MODE}
    except Exception as e:
        return {"ok": False, "error": str(e), "db_mode": DB_MODE}

@app.post("/auth/login")
def auth_login(payload: dict = Body(...)):
    uname = (payload.get("username") or "").strip()
    # self-heal: if the user table is empty (seed didn't run), create admin now
    try:
        if not DB.select("app_user", {"username": "admin"}):
            DB.insert("app_user", {"username": "admin", "pw_hash": _hash_pw("ChangeMe#2026"),
                                   "role": "admin", "must_change": True, "active": True,
                                   "created_at": datetime.datetime.utcnow().isoformat()})
    except Exception as e:
        raise HTTPException(500, f"User table not reachable: {e}")
    u = DB.select("app_user", {"username": uname})
    if not u or not u[0].get("active", True) or not _verify_pw(payload.get("password", ""), u[0]["pw_hash"]):
        raise HTTPException(401, "Invalid username or password")
    u = u[0]
    tok = _new_token()
    DB.insert("session", {"token": tok, "user_id": u["id"],
                          "created_at": datetime.datetime.utcnow().isoformat(),
                          "expires_at": (datetime.datetime.utcnow() + datetime.timedelta(days=7)).isoformat()})
    DB.update("app_user", {"id": u["id"]}, {"last_login": datetime.datetime.utcnow().isoformat()})
    return {"token": tok, "username": u["username"], "role": u["role"], "must_change": u.get("must_change", False)}

@app.get("/auth/me")
def auth_me(authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u: raise HTTPException(401, "not authenticated")
    return {"username": u["username"], "role": u["role"], "must_change": u.get("must_change", False)}

@app.post("/auth/change_password")
def auth_change(payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u: raise HTTPException(401, "not authenticated")
    newpw = payload.get("new_password", "")
    if len(newpw) < 8: raise HTTPException(400, "Password must be at least 8 characters")
    DB.update("app_user", {"id": u["id"]}, {"pw_hash": _hash_pw(newpw), "must_change": False})
    return {"ok": True}

@app.post("/auth/logout")
def auth_logout(authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    if tok: DB.delete("session", {"token": tok})
    return {"ok": True}

@app.get("/auth/users")
def auth_users(authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u or u["role"] != "admin": raise HTTPException(403, "admin only")
    return [{"id": x["id"], "username": x["username"], "role": x["role"],
             "active": x.get("active", True), "must_change": x.get("must_change", False),
             "last_login": x.get("last_login")} for x in DB.select("app_user", order="id")]

@app.post("/auth/users")
def auth_add_user(payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u or u["role"] != "admin": raise HTTPException(403, "admin only")
    uname = (payload.get("username") or "").strip()
    if not uname: raise HTTPException(400, "username required")
    if DB.select("app_user", {"username": uname}): raise HTTPException(400, "username exists")
    temp = payload.get("temp_password") or secrets.token_urlsafe(6)
    DB.insert("app_user", {"username": uname, "pw_hash": _hash_pw(temp),
                           "role": payload.get("role", "planner"), "must_change": True, "active": True,
                           "created_at": datetime.datetime.utcnow().isoformat()})
    return {"ok": True, "username": uname, "temp_password": temp}

@app.post("/auth/users/{uid}/reset")
def auth_reset_user(uid: int, authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u or u["role"] != "admin": raise HTTPException(403, "admin only")
    temp = secrets.token_urlsafe(6)
    DB.update("app_user", {"id": uid}, {"pw_hash": _hash_pw(temp), "must_change": True})
    return {"ok": True, "temp_password": temp}

@app.patch("/auth/users/{uid}")
def auth_patch_user(uid: int, payload: dict = Body(...), authorization: Optional[str] = Header(None)):
    tok = (authorization or "").replace("Bearer ", "")
    u = current_user(tok)
    if not u or u["role"] != "admin": raise HTTPException(403, "admin only")
    allowed = {k: v for k, v in payload.items() if k in ("role", "active")}
    if allowed: DB.update("app_user", {"id": uid}, allowed)
    return {"ok": True}

# ---- rate & dummy suggestions (Eleanor's logic) ----
@app.get("/rates/suggest")
def rates_suggest(version_id: int, months: int = 6):
    """Accessories % and Returns % of gross per DC from history (fallback current-cycle actuals).
    Eleanor: >=3-6 months of history, rate = returns / gross (and acc / gross)."""
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    hist = DB.select("history")
    src_label = "recency-weighted over all history"
    gross = defaultdict(float); ret = defaultdict(float); acc = defaultdict(float)
    if hist:
        latest = max((h["year"], h["month_no"]) for h in hist)
        for h in hist:
            age = (latest[0]*12+latest[1]) - (h["year"]*12+h["month_no"])
            if age < 0: continue
            w = 0.85 ** age
            if h["kind"] == "GROSS": gross[h["dc_code"]] += h["units"] * w
            elif h["kind"] == "RETURN": ret[h["dc_code"]] += h["units"] * w
        # accessories aren't in history table; use this week's actuals for acc ratio
    acts = DB.select("actual", {"version_id": version_id})
    cyc_gross = defaultdict(float)
    for a in acts:
        if a["kind"] == "GROSS": cyc_gross[a["dc_code"]] += a["units"]
        elif a["kind"] == "ACC": acc[a["dc_code"]] += a["units"]
        elif a["kind"] == "RETURN" and not hist: ret[a["dc_code"]] += a["units"]
    if not hist:
        src_label = "current cycle QTD actuals (import history for a real window)"
        gross = cyc_gross
    out = []
    for dc in ("ATL", "MX"):
        g = gross.get(dc, 0) or cyc_gross.get(dc, 0)
        gacc = cyc_gross.get(dc, 0)
        out.append({"dc_code": dc, "kind": "RET", "pct": round(ret.get(dc, 0) / g, 4) if g else 0})
        out.append({"dc_code": dc, "kind": "ACC", "pct": round(acc.get(dc, 0) / gacc, 4) if gacc else 0})
    return {"suggestions": out, "basis": src_label,
            "note": "Returns % from returns/gross; Accessories % from accessories/gross (QTD)."}

@app.get("/dummy_plan/suggest")
def dummy_suggest(version_id: int, weeks: int = 26):
    """Dummy run-rate from trailing history (Eleanor: ~800-1500/wk Atlanta average)."""
    v = DB.select("version", {"id": version_id})
    if not v: raise HTTPException(404, "version not found")
    hist = DB.select("history")
    # dummies live in actual table (kind DUMMY) for this week; use that as run-rate basis
    acts = DB.select("actual", {"version_id": version_id})
    by_dc_wk = defaultdict(list)
    for a in acts:
        if a["kind"] == "DUMMY": by_dc_wk[a["dc_code"]].append(a["units"])
    rr = {}
    for dc, vals in by_dc_wk.items():
        nz = [x for x in vals if x > 0]
        rr[dc] = round(sum(nz) / len(nz)) if nz else 0
    return {"run_rate": rr or {"ATL": 1200},
            "basis": "average of non-zero dummy weeks this cycle",
            "note": "Apply to forecast weeks; add supply-chain bulk drops manually in the known week."}

@app.get("/drilldown")
def drilldown(version_id: int):
    """To-Date / To-Go hierarchical drilldown (no year-over-year).
    TD = actualized (closed) weeks; TG = forecast (open) weeks. Flat facts; frontend builds the tree."""
    m = compute_model(version_id)
    periods = {p["id"]: p for p in m["periods"]}
    act_set = set(m["actualized_periods"])
    ch_names = {c["code"]: c["name"] for c in m["channels"]}
    MN = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    facts = []
    CNAME = {"IT": "Italy", "CN": "China", "TH": "Thailand", "": ""}
    def add(dc, country, mo, wk, product, channel, td, tg):
        if abs(td) < 0.5 and abs(tg) < 0.5: return
        facts.append({"half": ("H1" if mo <= 6 else "H2"), "quarter": f"Q{(mo-1)//3+1}",
                      "month": MN[mo], "month_no": mo, "week": f"W{wk}", "dc": dc,
                      "country": CNAME.get(country, country) or "Direct (RW/OW)",
                      "product": product, "channel": channel, "td": round(td), "tg": round(tg)})
    # ATL and MX: full channel + product detail, country not applicable
    for dc in ("ATL", "MX"):
        for r in m["grids"][dc]["rows"]:
            per = periods[r["period_id"]]; mo, wk = per["month_no"], per["week_no"]
            is_act = r["period_id"] in act_set
            for c in m["channels"]:
                v = r["cells"][c["code"]]["v"]
                add(dc, "", mo, wk, "Frames", ch_names[c["code"]], v if is_act else 0, 0 if is_act else v)
            for prod, val in [("Accessories", r["acc"]), ("Returns", r["ret"]), ("Dummies", r["dummy"]),
                              ("RW", r["rw"]), ("Nuance", r["nuance"]), ("OW", r["ow"])]:
                add(dc, "", mo, wk, prod, "-", val if is_act else 0, 0 if is_act else val)
    # DIRECT: the country split (IT/CN/TH) IS the Direct frames breakdown — emit under dc="DIRECT" with country.
    for country in ("IT", "CN", "TH"):
        for r in m["customer_grid"]["countries"][country]["rows"]:
            per = periods[r["period_id"]]; mo, wk = per["month_no"], per["week_no"]
            is_act = r["actual"]
            g = (r.get("actual_total") if is_act and r.get("actual_total") is not None
                 else r.get("reconciled_total", r["total"]))
            add("DIRECT", country, mo, wk, "Frames", "all", g if is_act else 0, 0 if is_act else g)
    # DIRECT meta lines (RW/OW) — not country-split, sit directly under DIRECT
    for r in m["grids"]["DIRECT"]["rows"]:
        per = periods[r["period_id"]]; mo, wk = per["month_no"], per["week_no"]
        is_act = r["period_id"] in act_set
        add("DIRECT", "", mo, wk, "Meta RW", "-", r.get("meta_rw", 0) if is_act else 0, 0 if is_act else r.get("meta_rw", 0))
        add("DIRECT", "", mo, wk, "Meta OW", "-", r.get("meta_ow", 0) if is_act else 0, 0 if is_act else r.get("meta_ow", 0))
    return {"facts": facts,
            "dimensions": ["half", "quarter", "month", "week", "dc", "country", "product", "channel"],
            "dim_labels": {"half": "Half", "quarter": "Quarter", "month": "Month", "week": "Week",
                           "dc": "DC", "country": "Country", "product": "Product", "channel": "Channel"}}

@app.post("/validate/legacy_export")
async def validate_legacy_export(file: UploadFile = File(...), version_id: int = Form(...),
                                 tolerance_pct: float = Form(2.0)):
    """Same as /validate/legacy but returns an Excel diff workbook (week x DC totals + channel detail)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    data = await file.read()
    # reuse legacy compare by calling internals
    m = compute_model(version_id)
    wb_in = load_wb(data)
    ch_codes = [c["code"] for c in m["channels"]]
    ch_names = {c["code"]: c["name"] for c in m["channels"]}
    legacy = _parse_legacy_channel_file(wb_in, m)
    out = openpyxl.Workbook(); out.remove(out.active)
    HDR = Font(bold=True, color="FFFFFF"); NAVY = PatternFill("solid", fgColor="1F3864")
    BAD = PatternFill("solid", fgColor="FCE4CC")
    def sheet_for(title, header, data_rows):
        ws = out.create_sheet(title)
        for j, h in enumerate(header, 1):
            c = ws.cell(1, j, h); c.font = HDR; c.fill = NAVY
        for i, row in enumerate(data_rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(i, j, val)
            # flag row if last col (status) is off
            if row[-1] == "off":
                for j in range(1, len(row) + 1): ws.cell(i, j).fill = BAD
        for j in range(1, len(header) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 14
        return ws
    if legacy:
        summ = []
        chan = []
        for dc, obj in legacy.items():
            app_rows = {r["label"]: r for r in m["grids"][dc]["rows"]}
            for (mo, wk), fr in obj["rows"].items():
                lbl = f"{mo}-{wk}"; ar = app_rows.get(lbl)
                if not ar: continue
                d = ar["gross"] - fr["total"]
                pct = (d / fr["total"] * 100) if fr["total"] else 0
                ok = abs(pct) <= tolerance_pct or abs(d) <= 25
                summ.append([dc, lbl, round(fr["total"]), round(ar["gross"]), round(d),
                             round(pct, 1), "ok" if ok else "off"])
                for i, ch in enumerate(ch_codes):
                    fv = fr["channels"][i]; av = ar["cells"][ch]["v"]; dd = av - fv
                    if abs(dd) > max(50, abs(fv) * tolerance_pct / 100):
                        chan.append([dc, lbl, ch_names[ch], round(fv), round(av), round(dd),
                                     "off"])
        sheet_for("Totals diff", ["DC", "Week", "File", "App", "Delta", "Delta%", "Status"], summ)
        sheet_for("Channel diffs", ["DC", "Week", "Channel", "File", "App", "Delta", "Status"], chan)
    else:
        sheet_for("Info", ["message"], [["Upload an Atlanta/Mexico legacy workbook for channel-level export"]])
    buf = io.BytesIO(); out.save(buf); buf.seek(0)
    fn = f"Validation_diff_{m['version']['week_tag']}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fn}"'})

# serve the frontend when web/index.html sits next to app.py (single-host mode)
_WEB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "web")
if os.path.exists(os.path.join(_WEB, "index.html")):
    from fastapi.responses import FileResponse
    @app.get("/")
    def _index():
        return FileResponse(os.path.join(_WEB, "index.html"))


@app.patch("/admin/period/{pid}")
def patch_period(pid: int, payload: dict = Body(...)):
    if "working_days" in payload:
        DB.update("period", {"id": pid}, {"working_days": max(1, min(7, int(payload["working_days"])))})
    return {"ok": True}


# ------------------------------------------------------------------
# validation — source tie-outs + legacy parallel-run diff
# ------------------------------------------------------------------
@app.get("/import/log")
def import_log_list():
    rows = DB.select("import_log", order="-id")
    return rows[:40]

@app.get("/validate/tieout")
def validate_tieout(version_id: int):
    """Automatic checks: app values vs the source files as imported, plus internal cross-foots."""
    m = compute_model(version_id)
    checks = []
    def add(name, app_val, file_val, tol=1.0, note=""):
        app_r, file_r = round(app_val), (round(file_val) if file_val is not None else None)
        ok = file_r is None or abs(app_r - file_r) <= tol
        checks.append({"name": name, "app": app_r, "file": file_r,
                       "delta": (app_r - file_r) if file_r is not None else None,
                       "ok": bool(ok), "note": note})
    logs = DB.select("import_log", order="-id")
    def last_log(kind):
        for l in logs:
            if l["kind"] == kind: return l
        return None

    # 1) NA tie-out: forecast_channel in DB vs NA import preview totals
    na = last_log("NA")
    fc_rows = DB.select("forecast_channel", {"version_id": version_id})
    fc_tot = defaultdict(float)
    for r in fc_rows: fc_tot[r["channel_code"]] += r["units"]
    if na and na.get("meta", {}).get("totals"):
        for ch, v in na["meta"]["totals"].items():
            add(f"NA file · {ch}", fc_tot.get(ch, 0), v, note="forecast in app vs NA import")
        add("NA file · grand total (frames)",
            sum(fc_tot[c["code"]] for c in m["channels"]), na["meta"].get("grand_total"),
            note="sum of 10 frame channels")
    else:
        checks.append({"name": "NA file", "app": None, "file": None, "delta": None, "ok": False,
                       "note": "no NA import logged yet"})

    # 2) Actuals tie-out: DB gross by DC vs actuals import preview
    act_log = last_log("ACTUALS")
    acts = DB.select("actual", {"version_id": version_id})
    by_dc = defaultdict(float)
    for a in acts:
        if a["kind"] == "GROSS" and a["channel_code"]: by_dc[a["dc_code"]] += a["units"]
    if act_log and act_log.get("meta", {}).get("gross_by_dc"):
        for dc, v in act_log["meta"]["gross_by_dc"].items():
            add(f"Actuals file · {dc} net units", by_dc.get(dc, 0), v, note="QTD net in app vs BO import")
    # 3) Accessories tie-out
    acc_log = last_log("ACC")
    acc_dc = defaultdict(float)
    for a in acts:
        if a["kind"] == "ACC": acc_dc[a["dc_code"]] += a["units"]
    if acc_log and acc_log.get("meta", {}).get("by_dc"):
        for dc, v in acc_log["meta"]["by_dc"].items():
            add(f"Accessories file · {dc}", acc_dc.get(dc, 0), v)

    # 4) internal cross-foot per channel: ATL + MX + Direct vs NA forecast (forecast weeks)
    act_set = set(m["actualized_periods"])
    fc_by = defaultdict(float)
    for r in fc_rows:
        if r["period_id"] not in act_set:
            fc_by[r["channel_code"]] += r["units"]
    for c in m["channels"]:
        ch = c["code"]
        alloc = sum(r["cells"][ch]["v"] for dc in ("ATL", "MX", "DIRECT")
                    for r in m["grids"][dc]["rows"] if r["period_id"] not in act_set)
        add(f"Cross-foot · {ch}", alloc, fc_by.get(ch, 0), tol=2,
            note="ATL+MX+Direct forecast weeks vs NA — a gap here = overrides (intentional)")

    # 5) recon + negatives + actualization status
    rec = m.get("direct_recon", {})
    summary = {"actualized_weeks": [p["label"] for p in m["periods"]
                                    if p["id"] in act_set],
               "warnings": len(m["warnings"]),
               "direct_recon": rec,
               "passed": sum(1 for c in checks if c["ok"]),
               "failed": sum(1 for c in checks if not c["ok"])}
    return {"checks": checks, "summary": summary}

def _parse_legacy_channel_file(wb, m):
    """Parse Eleanor-format Atlanta/Mexico/Direct tabs: A=week B=month C=total D..M=10 channels."""
    out = {}
    name_map = {"ATL": "Atlanta", "MX": "Mexico", "DIRECT": "Direct"}
    for dc, key in name_map.items():
        sheet = next((s for s in wb.sheetnames if s.lower().startswith(key.lower())), None)
        if not sheet: continue
        ws = wb[sheet]
        rows = {}
        for r in range(3, 60):
            wk, mo = ws.cell(r, 1).value, ws.cell(r, 2).value
            if wk is None or mo is None:
                if r > 6: break
                continue
            try: k = (int(mo), int(wk))
            except Exception: continue
            rows[k] = {"total": num(ws.cell(r, 3).value),
                       "channels": [num(ws.cell(r, j).value) for j in range(4, 14)]}
        out[dc] = {"sheet": sheet, "rows": rows}
    return out

@app.post("/validate/legacy")
async def validate_legacy(file: UploadFile = File(...), version_id: int = Form(...),
                          tolerance_pct: float = Form(2.0)):
    """Parallel-run diff: upload the legacy Atlanta/Mexico (or Direct-country) workbook and
    compare week x DC totals — and per-channel — against the app model. Eyeball layer."""
    data = await file.read()
    m = compute_model(version_id)
    wb = load_wb(data)
    ch_codes = [c["code"] for c in m["channels"]]
    ch_names = {c["code"]: c["name"] for c in m["channels"]}
    legacy = _parse_legacy_channel_file(wb, m)
    if not legacy:
        # try Direct-country file: Italy/China/Thailand tabs, col C total
        out_c = {}
        for dc, key in [("IT", "Italy"), ("CN", "China"), ("TH", "Thailand")]:
            sheet = next((s for s in wb.sheetnames if s.lower().startswith(key.lower())), None)
            if not sheet: continue
            ws = wb[sheet]
            # locate Week/Month columns by header text (layouts differ between files)
            wk_c = mo_c = None
            for hr in (1, 2, 3):
                for j in range(1, 8):
                    v2 = str(ws.cell(hr, j).value or "").strip().lower()
                    if v2 == "week": wk_c = j
                    if v2 == "month": mo_c = j
                if wk_c and mo_c: break
            if not (wk_c and mo_c): wk_c, mo_c = 1, 2
            tot_c = mo_c + 1
            rows = {}
            for r in range(3, 60):
                wk, mo = ws.cell(r, wk_c).value, ws.cell(r, mo_c).value
                if wk is None or mo is None:
                    if r > 6: break
                    continue
                try: rows[(int(mo), int(wk))] = num(ws.cell(r, tot_c).value)
                except Exception: continue
            out_c[dc] = {"sheet": sheet, "rows": rows}
        if not out_c:
            raise HTTPException(400, "Could not find Atlanta/Mexico/Direct or Italy/China/Thailand tabs")
        diffs = []
        for dc, obj in out_c.items():
            app_rows = {r["label"]: (r.get("actual_total") if r["actual"] and r.get("actual_total") is not None
                                     else r.get("reconciled_total", r["total"]))
                        for r in m["customer_grid"]["countries"][dc]["rows"]}
            for (mo, wk), fv in obj["rows"].items():
                lbl = f"{mo}-{wk}"
                av = app_rows.get(lbl)
                if av is None: continue
                d = av - fv
                pct_d = (d / fv * 100) if fv else (0 if abs(d) < 1 else 100)
                diffs.append({"scope": dc, "week": lbl, "file": round(fv), "app": round(av),
                              "delta": round(d), "pct": round(pct_d, 1),
                              "ok": abs(pct_d) <= tolerance_pct or abs(d) <= 25})
        return {"kind": "direct_countries", "tolerance_pct": tolerance_pct, "rows": diffs,
                "passed": sum(1 for d in diffs if d["ok"]), "failed": sum(1 for d in diffs if not d["ok"])}

    diffs, chan_diffs = [], []
    for dc, obj in legacy.items():
        app_rows = {r["label"]: r for r in m["grids"][dc]["rows"]}
        for (mo, wk), fr in obj["rows"].items():
            lbl = f"{mo}-{wk}"
            ar = app_rows.get(lbl)
            if ar is None: continue
            d = ar["gross"] - fr["total"]
            pct_d = (d / fr["total"] * 100) if fr["total"] else (0 if abs(d) < 1 else 100)
            # per-channel breakdown for THIS week (for inline expand)
            ch_break = []
            for i, ch in enumerate(ch_codes):
                fv = fr["channels"][i]; av = ar["cells"][ch]["v"]; dd = av - fv
                ch_break.append({"channel": ch_names[ch], "file": round(fv), "app": round(av), "delta": round(dd)})
                if abs(dd) > max(50, abs(fv) * tolerance_pct / 100):
                    chan_diffs.append({"scope": dc, "week": lbl, "channel": ch_names[ch],
                                       "file": round(fv), "app": round(av), "delta": round(dd)})
            ch_break.sort(key=lambda x: -abs(x["delta"]))
            diffs.append({"scope": dc, "week": lbl, "actual": ar["actual"],
                          "file": round(fr["total"]), "app": round(ar["gross"]),
                          "delta": round(d), "pct": round(pct_d, 1),
                          "ok": abs(pct_d) <= tolerance_pct or abs(d) <= 25,
                          "channels": ch_break})
    chan_diffs.sort(key=lambda x: -abs(x["delta"]))
    return {"kind": "channel_file", "tolerance_pct": tolerance_pct, "rows": diffs,
            "channel_diffs": chan_diffs[:30],
            "passed": sum(1 for d in diffs if d["ok"]), "failed": sum(1 for d in diffs if not d["ok"])}
